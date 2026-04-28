"""Тести для checkr.py — автоматична валідація товарного фіду e-commerce.

Запуск:
    pytest tests/test_checkr.py -v
"""

import os
import tempfile
from pathlib import Path

import pandas as pd
import pytest

from checkr import (
    VALIDATION_RULES,
    battery_to_mah,
    check_conflicts,
    check_semantic_conflicts,
    clean_html,
    extract_battery_matches,
    extract_memory_matches,
    extract_memory_values,
    extract_resolution_matches,
    extract_screen_diagonal_matches,
    extract_value_list_matches,
    extract_weight_matches,
    find_column,
    memory_to_mb,
    normalize_battery_value,
    normalize_memory_value,
    read_input_file,
    validate_feed,
    weight_to_grams,
)


# ===========================================================================
# Тести Модуля 1: clean_html
# ===========================================================================


class TestCleanHtml:
    def test_removes_paragraph_tags(self):
        result = clean_html("<p>Hello <b>World</b></p>")
        assert "Hello" in result
        assert "World" in result
        assert "<" not in result

    def test_removes_all_html_tags(self):
        result = clean_html("<div><h1>Title</h1><p>Body</p></div>")
        assert "<" not in result
        assert "Title" in result
        assert "Body" in result

    def test_returns_empty_for_none(self):
        assert clean_html(None) == ""

    def test_returns_empty_for_integer(self):
        assert clean_html(42) == ""

    def test_returns_empty_for_float(self):
        assert clean_html(3.14) == ""

    def test_returns_unchanged_plain_text(self):
        assert clean_html("plain text") == "plain text"

    def test_returns_empty_for_empty_string(self):
        # Порожній рядок повертається як є (без змін)
        assert clean_html("") == ""

    def test_extracts_text_from_strong_tag(self):
        result = clean_html("Ноутбук із <strong>512 ГБ</strong> SSD")
        assert "512 ГБ" in result

    def test_handles_nested_tags(self):
        html = "<ul><li>512 ГБ SSD</li><li>16 GB RAM</li></ul>"
        result = clean_html(html)
        assert "512 ГБ SSD" in result
        assert "16 GB RAM" in result


# ===========================================================================
# Тести Модуля 2: normalize_memory_value
# ===========================================================================


class TestNormalizeMemoryValue:
    # Гігабайти: різні варіанти написання
    def test_gb_latin_uppercase(self):
        assert normalize_memory_value("512 GB") == "512ГБ"

    def test_gb_latin_no_space(self):
        assert normalize_memory_value("512GB") == "512ГБ"

    def test_gb_cyrillic_uppercase(self):
        assert normalize_memory_value("512 ГБ") == "512ГБ"

    def test_gb_cyrillic_no_space(self):
        assert normalize_memory_value("512ГБ") == "512ГБ"

    def test_gib(self):
        assert normalize_memory_value("512 GiB") == "512ГБ"

    # Терабайти: різні варіанти написання
    def test_tb_latin_uppercase(self):
        assert normalize_memory_value("1 TB") == "1ТБ"

    def test_tb_latin_no_space(self):
        assert normalize_memory_value("1TB") == "1ТБ"

    def test_tb_cyrillic_uppercase(self):
        assert normalize_memory_value("1 ТБ") == "1ТБ"

    def test_tb_cyrillic_no_space(self):
        assert normalize_memory_value("1ТБ") == "1ТБ"

    def test_tib(self):
        assert normalize_memory_value("2 TiB") == "2ТБ"

    # Мегабайти
    def test_mb_latin(self):
        assert normalize_memory_value("256 MB") == "256МБ"

    def test_mb_cyrillic(self):
        assert normalize_memory_value("256МБ") == "256МБ"

    def test_mib(self):
        assert normalize_memory_value("256 MiB") == "256МБ"

    # Дробові значення
    def test_decimal_dot(self):
        assert normalize_memory_value("1.5 TB") == "1.5ТБ"

    def test_decimal_comma(self):
        assert normalize_memory_value("2,5 ГБ") == "2.5ГБ"

    def test_decimal_trailing_zero_removed(self):
        assert normalize_memory_value("1.0 TB") == "1ТБ"

    def test_decimal_trailing_zeros_removed(self):
        assert normalize_memory_value("1.50 TB") == "1.5ТБ"

    # Регістр
    def test_lowercase_gb(self):
        assert normalize_memory_value("512 gb") == "512ГБ"

    def test_lowercase_tb(self):
        assert normalize_memory_value("1 tb") == "1ТБ"

    # Невідоме значення
    def test_unknown_returns_uppercase(self):
        result = normalize_memory_value("unknown")
        assert result == "UNKNOWN"


# ===========================================================================
# Тести Модуля 2: extract_memory_matches та extract_memory_values
# ===========================================================================


class TestExtractMemoryMatches:
    def test_single_value(self):
        matches = extract_memory_matches("Ноутбук з 512 ГБ SSD")
        assert len(matches) == 1
        raw, norm = matches[0]
        assert norm == "512ГБ"

    def test_multiple_values(self):
        matches = extract_memory_matches("512 ГБ SSD та 32 GB RAM")
        norms = [n for _, n in matches]
        assert "512ГБ" in norms
        assert "32ГБ" in norms

    def test_empty_string(self):
        assert extract_memory_matches("") == []

    def test_none_input(self):
        assert extract_memory_matches(None) == []

    def test_whitespace_only(self):
        assert extract_memory_matches("   ") == []

    def test_no_memory_values(self):
        assert extract_memory_matches("Просто текст без пам'яті") == []

    def test_returns_original_and_normalized(self):
        matches = extract_memory_matches("512GB")
        assert len(matches) == 1
        raw, norm = matches[0]
        assert norm == "512ГБ"

    def test_does_not_match_partial_number(self):
        # "51024GB" не повинно збігатися як "1024GB" після lookbehind
        matches = extract_memory_matches("51024GB")
        # Може збігатися як "51024GB" або не збігатися зовсім — головне, не "1024GB"
        norms = [n for _, n in matches]
        assert "1024ГБ" not in norms


class TestExtractMemoryValues:
    def test_returns_only_normalized(self):
        result = extract_memory_values("Ноутбук 512 ГБ SSD та 32 GB RAM")
        assert result == ["512ГБ", "32ГБ"]

    def test_empty_input(self):
        assert extract_memory_values("") == []

    def test_none_input(self):
        assert extract_memory_values(None) == []

    def test_no_match(self):
        assert extract_memory_values("Просто текст") == []


# ===========================================================================
# Тести Модуля 2b: extract_screen_diagonal_matches
# ===========================================================================


class TestExtractScreenDiagonalMatches:
    def test_inch_symbol(self):
        matches = extract_screen_diagonal_matches('Ноутбук 15.6" IPS')
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == '15.6"'

    def test_unicode_inch(self):
        matches = extract_screen_diagonal_matches("Дисплей 14″")
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == '14"'

    def test_dyuim_word(self):
        matches = extract_screen_diagonal_matches("Екран 13.3 дюймів")
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == '13.3"'

    def test_inch_word(self):
        matches = extract_screen_diagonal_matches("15.6-inch display")
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == '15.6"'

    def test_comma_decimal(self):
        matches = extract_screen_diagonal_matches('15,6"')
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == '15.6"'

    def test_integer_diagonal(self):
        matches = extract_screen_diagonal_matches('14" ноутбук')
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == '14"'

    def test_multiple_diagonals(self):
        matches = extract_screen_diagonal_matches('15.6" і 14" дисплеї')
        norms = [n for _, n in matches]
        assert '15.6"' in norms
        assert '14"' in norms

    def test_no_unit_no_match(self):
        # Без одиниці виміру не має збігу (уникаємо хибних спрацювань)
        assert extract_screen_diagonal_matches("Ноутбук 15.6 IPS") == []

    def test_empty_string(self):
        assert extract_screen_diagonal_matches("") == []

    def test_none_input(self):
        assert extract_screen_diagonal_matches(None) == []

    def test_trailing_zero_removed(self):
        matches = extract_screen_diagonal_matches('15.60"')
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == '15.6"'


# ===========================================================================
# Тести Модуля 2b: extract_weight_matches
# ===========================================================================


class TestExtractWeightMatches:
    def test_kg_cyrillic(self):
        matches = extract_weight_matches("Вага: 1.5 кг")
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == "1.5кг"

    def test_kg_latin(self):
        matches = extract_weight_matches("Weight: 2.3kg")
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == "2.3кг"

    def test_comma_decimal(self):
        matches = extract_weight_matches("1,8 кг")
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == "1.8кг"

    def test_integer_weight(self):
        matches = extract_weight_matches("2 кг")
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == "2кг"

    def test_no_match_without_unit(self):
        assert extract_weight_matches("Ноутбук 1.5") == []

    def test_does_not_match_gb(self):
        # "GB" не має сприйматися як вага
        assert extract_weight_matches("512 GB SSD") == []

    def test_empty_string(self):
        assert extract_weight_matches("") == []

    def test_none_input(self):
        assert extract_weight_matches(None) == []

    def test_trailing_zero_removed(self):
        matches = extract_weight_matches("1.50 кг")
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == "1.5кг"


# ===========================================================================
# Тести Модуля 2b: extract_resolution_matches
# ===========================================================================


class TestExtractResolutionMatches:
    def test_numerical_x(self):
        matches = extract_resolution_matches("Дисплей 1920x1080")
        norms = [n for _, n in matches]
        assert "1920x1080" in norms

    def test_numerical_unicode_x(self):
        matches = extract_resolution_matches("3840×2160")
        norms = [n for _, n in matches]
        assert "3840x2160" in norms

    def test_fhd_alias(self):
        matches = extract_resolution_matches("Екран FHD")
        norms = [n for _, n in matches]
        assert "1920x1080" in norms

    def test_full_hd_alias(self):
        matches = extract_resolution_matches("Full HD дисплей")
        norms = [n for _, n in matches]
        assert "1920x1080" in norms

    def test_4k_alias(self):
        matches = extract_resolution_matches("Монітор 4K")
        norms = [n for _, n in matches]
        assert "3840x2160" in norms

    def test_qhd_alias(self):
        matches = extract_resolution_matches("QHD панель")
        norms = [n for _, n in matches]
        assert "2560x1440" in norms

    def test_hd_alias(self):
        matches = extract_resolution_matches("HD екран")
        norms = [n for _, n in matches]
        assert "1366x768" in norms

    def test_full_hd_preferred_over_hd(self):
        # "Full HD" не має повертати окремий збіг для "HD"
        matches = extract_resolution_matches("Full HD дисплей")
        norms = [n for _, n in matches]
        assert "1920x1080" in norms
        assert norms.count("1366x768") == 0

    def test_empty_string(self):
        assert extract_resolution_matches("") == []

    def test_none_input(self):
        assert extract_resolution_matches(None) == []

    def test_no_resolution_in_text(self):
        assert extract_resolution_matches("Просто текст") == []

    def test_hdmi_not_matched_as_hd(self):
        """HDMI не має розпізнаватись як HD (word boundary check)."""
        matches = extract_resolution_matches("Підтримка HDMI та DisplayPort")
        # Не повинно бути жодних збігів з роздільною здатністю
        assert matches == []
    
    def test_hd_in_sentence_with_hdmi(self):
        """HD має знаходитись окремо, але не в HDMI."""
        matches = extract_resolution_matches("Екран HD з підтримкою HDMI")
        norms = [n for _, n in matches]
        # Має бути один збіг для "HD", але не для "HDMI"
        assert "1366x768" in norms
        assert len(matches) == 1


# ===========================================================================
# Тести Модуля 2b: extract_value_list_matches
# ===========================================================================


class TestExtractValueListMatches:
    MATRIX_TYPES = ["IPS", "TN", "VA", "OLED", "AMOLED"]

    def test_finds_ips(self):
        matches = extract_value_list_matches("Матриця IPS, тонкі рамки", self.MATRIX_TYPES)
        norms = [n for _, n in matches]
        assert "IPS" in norms

    def test_case_insensitive(self):
        matches = extract_value_list_matches("ips матриця", self.MATRIX_TYPES)
        assert len(matches) == 1
        _, norm = matches[0]
        assert norm == "IPS"

    def test_finds_oled(self):
        matches = extract_value_list_matches("OLED дисплей", self.MATRIX_TYPES)
        norms = [n for _, n in matches]
        assert "OLED" in norms

    def test_amoled_before_oled(self):
        # "AMOLED" довший за "OLED" — має знаходитися першим, не дублювати "OLED"
        matches = extract_value_list_matches("AMOLED екран", self.MATRIX_TYPES)
        norms = [n for _, n in matches]
        assert "AMOLED" in norms

    def test_no_match(self):
        assert extract_value_list_matches("Звичайний текст", self.MATRIX_TYPES) == []

    def test_empty_string(self):
        assert extract_value_list_matches("", self.MATRIX_TYPES) == []

    def test_empty_valid_values(self):
        assert extract_value_list_matches("IPS матриця", []) == []

    def test_none_input(self):
        assert extract_value_list_matches(None, self.MATRIX_TYPES) == []





class TestFindColumn:
    COLS = ["Название", "Назва (ua)", "Объём SSD;115411",
            "Объем установленной оперативной памяти;20863", "Описание;1",
            "Краткое описание"]

    def test_exact_match(self):
        assert find_column(self.COLS, "Название") == "Название"

    def test_partial_match_ssd(self):
        assert find_column(self.COLS, "Объём SSD") == "Объём SSD;115411"

    def test_partial_match_ram(self):
        result = find_column(self.COLS, "Объем установленной оперативной памяти")
        assert result == "Объем установленной оперативной памяти;20863"

    def test_finds_opisanie_not_kratkoe(self):
        # "Описание" має знаходити "Описание;1", а НЕ "Краткое описание"
        assert find_column(self.COLS, "Описание") == "Описание;1"

    def test_finds_nazva_ua_not_nazvanie(self):
        # "Назва" має знаходити "Назва (ua)", а НЕ "Название"
        assert find_column(self.COLS, "Назва") == "Назва (ua)"

    def test_not_found(self):
        assert find_column(self.COLS, "Відеокарта") is None

    def test_case_insensitive_exact(self):
        assert find_column(self.COLS, "название") == "Название"

    def test_case_insensitive_partial(self):
        assert find_column(self.COLS, "объём ssd") == "Объём SSD;115411"

    def test_empty_columns(self):
        assert find_column([], "Назва") is None

    def test_prefers_exact_over_partial(self):
        cols = ["SSD", "Объём SSD;115411"]
        # "SSD" — точний збіг, тому його повернути першим
        assert find_column(cols, "SSD") == "SSD"

    def test_ssd_found_via_content_match(self):
        # hint="SSD" знаходить "Объём SSD;115411" через часткове входження у base
        cols = ["Объём SSD;115411"]
        assert find_column(cols, "SSD") == "Объём SSD;115411"


# ===========================================================================
# Тести Модуля 5: check_conflicts
# ===========================================================================


class TestCheckConflicts:
    """Тести логіки виявлення конфліктів."""

    def _ssd_rule(self) -> dict:
        return next(r for r in VALIDATION_RULES if r["label"] == "SSD")

    def _ram_rule(self) -> dict:
        return next(r for r in VALIDATION_RULES if r["label"] == "RAM")

    # --- Тести без конфлікту ---

    def test_no_conflict_matching_values(self):
        row = pd.Series({
            "Название": "Ноутбук 512 ГБ SSD",
            "Краткое описание": "Ноутбук із 512 ГБ SSD",
            "Объём SSD;115411": "512 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error == ""
        assert cols == []

    def test_no_conflict_text_has_no_memory_value(self):
        # Текст не згадує пам'ять — це не конфлікт
        row = pd.Series({
            "Название": "Ноутбук ASUS VivoBook",
            "Краткое описание": "Потужний ноутбук",
            "Объём SSD;115411": "512 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error == ""
        assert cols == []

    def test_no_conflict_empty_characteristic(self):
        row = pd.Series({
            "Название": "Ноутбук 512 ГБ SSD",
            "Объём SSD;115411": "",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error == ""
        assert cols == []

    def test_no_conflict_missing_char_column(self):
        # Колонка характеристики відсутня — не конфлікт
        row = pd.Series({"Название": "Ноутбук 512 ГБ SSD"})
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error == ""
        assert cols == []

    def test_no_conflict_nan_characteristic(self):
        row = pd.Series({
            "Название": "Ноутбук 512 ГБ SSD",
            "Объём SSD;115411": float("nan"),
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error == ""
        assert cols == []

    # --- Тести з конфліктом ---

    def test_conflict_in_name(self):
        # Назва: 128ГБ, Характеристика: 512ГБ
        row = pd.Series({
            "Название": "Ноутбук 128 ГБ SSD",
            "Объём SSD;115411": "512 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error != ""
        assert "128ГБ" in error or "128 ГБ" in error
        assert "512ГБ" in error or "512 ГБ" in error
        assert "Название" in cols
        assert "Объём SSD;115411" in cols

    def test_conflict_in_description(self):
        # Опис: 256ГБ, Характеристика: 512ГБ
        row = pd.Series({
            "Краткое описание": "Ноутбук із 256 ГБ SSD",
            "Объём SSD;115411": "512 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error != ""
        assert "256ГБ" in error or "256 ГБ" in error

    def test_conflict_in_html_description(self):
        # Опис з HTML-тегами: виявити конфлікт після очищення
        row = pd.Series({
            "Описание;1": "<p>Ноутбук із <b>128 ГБ</b> SSD</p>",
            "Объём SSD;115411": "512 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error != ""
        assert "Конфлікт SSD" in error

    def test_conflict_mixed_notation(self):
        # Назва: "512 GB" (латиниця), Характеристика: "512 ГБ" (кирилиця) — НЕ конфлікт
        row = pd.Series({
            "Название": "Ноутбук 512 GB SSD",
            "Объём SSD;115411": "512 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        # Обидва нормалізуються до "512ГБ" — конфлікту немає
        assert error == ""

    def test_conflict_error_message_format(self):
        row = pd.Series({
            "Название": "Ноутбук 128 ГБ SSD",
            "Объём SSD;115411": "512 ГБ",
        })
        error, _ = check_conflicts(row, list(row.index), self._ssd_rule())
        assert "Конфлікт SSD" in error
        assert "Характеристика" in error

    def test_conflict_returns_all_conflicting_columns(self):
        # Конфліктні значення в обох текстових полях
        row = pd.Series({
            "Название": "Ноутбук 128 ГБ SSD",
            "Краткое описание": "Ноутбук із 256 ГБ SSD",
            "Объём SSD;115411": "512 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert "Название" in cols
        assert "Краткое описание" in cols
        assert "Объём SSD;115411" in cols

    def test_conflict_ram_rule(self):
        row = pd.Series({
            "Название": "Ноутбук 4 GB RAM",
            "Объем установленной оперативной памяти;20863": "32 GB",
        })
        error, cols = check_conflicts(row, list(row.index), self._ram_rule())
        assert error != ""
        assert "Конфлікт RAM" in error


# ===========================================================================
# Тести Модуля 5: check_conflicts — нові типи перевірок
# ===========================================================================


class TestCheckConflictsNewTypes:
    """Тести логіки виявлення конфліктів для нових checker_type."""

    # --- screen_diagonal ---

    def _diagonal_rule(self) -> dict:
        return {
            "label": "Діагональ екрана",
            "checker_type": "screen_diagonal",
            "char_hints": ["Діагональ"],
            "text_hints": ["Название"],
        }

    def test_diagonal_no_conflict_matching(self):
        row = pd.Series({
            "Название": 'Ноутбук 15.6" IPS',
            "Діагональ": "15.6",
        })
        error, cols = check_conflicts(row, list(row.index), self._diagonal_rule())
        assert error == ""

    def test_diagonal_conflict(self):
        row = pd.Series({
            "Название": '14" ноутбук',
            "Діагональ": "15.6",
        })
        error, cols = check_conflicts(row, list(row.index), self._diagonal_rule())
        assert error != ""
        assert "Конфлікт Діагональ екрана" in error
        assert "Діагональ" in cols

    def test_diagonal_no_mention_in_text(self):
        row = pd.Series({
            "Название": "Ноутбук ASUS VivoBook",
            "Діагональ": "15.6",
        })
        error, cols = check_conflicts(row, list(row.index), self._diagonal_rule())
        assert error == ""

    def test_diagonal_char_col_missing(self):
        row = pd.Series({"Название": '15.6" ноутбук'})
        error, cols = check_conflicts(row, list(row.index), self._diagonal_rule())
        assert error == ""

    # --- weight ---

    def _weight_rule(self) -> dict:
        return {
            "label": "Вага",
            "checker_type": "weight",
            "char_hints": ["Вага"],
            "text_hints": ["Название"],
        }

    def test_weight_no_conflict(self):
        row = pd.Series({
            "Название": "Ноутбук вагою 1.5 кг",
            "Вага": "1.5 кг",
        })
        error, cols = check_conflicts(row, list(row.index), self._weight_rule())
        assert error == ""

    def test_weight_conflict(self):
        row = pd.Series({
            "Название": "Ноутбук вагою 2.3 кг",
            "Вага": "1.5 кг",
        })
        error, cols = check_conflicts(row, list(row.index), self._weight_rule())
        assert error != ""
        assert "Конфлікт Вага" in error

    def test_weight_bare_number_in_char_col(self):
        # Колонка характеристики містить просте число (без "кг")
        row = pd.Series({
            "Название": "Ноутбук вагою 1.5 кг",
            "Вага": "1.5",
        })
        error, cols = check_conflicts(row, list(row.index), self._weight_rule())
        assert error == ""

    # --- resolution ---

    def _resolution_rule(self) -> dict:
        return {
            "label": "Роздільна здатність",
            "checker_type": "resolution",
            "char_hints": ["Розд"],
            "text_hints": ["Название"],
        }

    def test_resolution_no_conflict_fhd(self):
        row = pd.Series({
            "Название": "Ноутбук FHD дисплей",
            "Розд": "1920x1080",
        })
        error, cols = check_conflicts(row, list(row.index), self._resolution_rule())
        assert error == ""

    def test_resolution_conflict(self):
        row = pd.Series({
            "Название": "Ноутбук HD екран",
            "Розд": "1920x1080",
        })
        error, cols = check_conflicts(row, list(row.index), self._resolution_rule())
        assert error != ""
        assert "Конфлікт Роздільна здатність" in error

    def test_resolution_alias_in_char_col(self):
        # Колонка характеристики містить псевдонім "FHD"
        row = pd.Series({
            "Название": "Ноутбук 1920x1080",
            "Розд": "FHD",
        })
        error, cols = check_conflicts(row, list(row.index), self._resolution_rule())
        assert error == ""

    # --- value_list ---

    def _matrix_rule(self) -> dict:
        return {
            "label": "Тип матриці",
            "checker_type": "value_list",
            "char_hints": ["Матриця"],
            "text_hints": ["Название"],
            "valid_values": ["IPS", "TN", "VA", "OLED"],
        }

    def test_value_list_no_conflict(self):
        row = pd.Series({
            "Название": "Ноутбук IPS матриця",
            "Матриця": "IPS",
        })
        error, cols = check_conflicts(row, list(row.index), self._matrix_rule())
        assert error == ""

    def test_value_list_conflict(self):
        row = pd.Series({
            "Название": "Ноутбук TN матриця",
            "Матриця": "IPS",
        })
        error, cols = check_conflicts(row, list(row.index), self._matrix_rule())
        assert error != ""
        assert "Конфлікт Тип матриці" in error

    def test_value_list_no_mention(self):
        row = pd.Series({
            "Название": "Ноутбук без згадки типу",
            "Матриця": "IPS",
        })
        error, cols = check_conflicts(row, list(row.index), self._matrix_rule())
        assert error == ""

    def test_value_list_case_insensitive(self):
        row = pd.Series({
            "Название": "Ноутбук ips матриця",
            "Матриця": "IPS",
        })
        error, cols = check_conflicts(row, list(row.index), self._matrix_rule())
        assert error == ""


# ===========================================================================
# Тести validation_rules.py
# ===========================================================================


class TestValidationRules:
    """Тести структури VALIDATION_RULES."""

    def test_has_ssd_rule(self):
        labels = [r["label"] for r in VALIDATION_RULES]
        assert "SSD" in labels

    def test_has_ram_rule(self):
        labels = [r["label"] for r in VALIDATION_RULES]
        assert "RAM" in labels

    def test_has_screen_diagonal_rule(self):
        labels = [r["label"] for r in VALIDATION_RULES]
        assert "Діагональ екрана" in labels

    def test_has_weight_rule(self):
        labels = [r["label"] for r in VALIDATION_RULES]
        assert "Вага" in labels

    def test_has_resolution_rule(self):
        labels = [r["label"] for r in VALIDATION_RULES]
        assert "Роздільна здатність" in labels

    def test_has_matrix_type_rule(self):
        labels = [r["label"] for r in VALIDATION_RULES]
        assert "Тип матриці" in labels

    def test_has_gpu_type_rule(self):
        labels = [r["label"] for r in VALIDATION_RULES]
        assert "Тип відеокарти" in labels

    def test_all_rules_have_required_fields(self):
        for rule in VALIDATION_RULES:
            assert "label" in rule, f"Rule missing 'label': {rule}"
            assert "checker_type" in rule, f"Rule missing 'checker_type': {rule}"
            assert "char_hints" in rule, f"Rule missing 'char_hints': {rule}"
            assert "text_hints" in rule, f"Rule missing 'text_hints': {rule}"

    def test_value_list_rules_have_valid_values(self):
        for rule in VALIDATION_RULES:
            if rule.get("checker_type") == "value_list":
                assert "valid_values" in rule, (
                    f"value_list rule missing 'valid_values': {rule['label']}"
                )
                assert len(rule["valid_values"]) > 0

    def test_memory_rules_have_memory_checker_type(self):
        ssd_rule = next(r for r in VALIDATION_RULES if r["label"] == "SSD")
        ram_rule = next(r for r in VALIDATION_RULES if r["label"] == "RAM")
        assert ssd_rule["checker_type"] == "memory"
        assert ram_rule["checker_type"] == "memory"





class TestReadInputFile:
    def test_reads_csv_with_comma_separator(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text(
            "Название,Объём SSD;115411\n"
            "Ноутбук 512 ГБ,512 ГБ\n",
            encoding="utf-8"
        )
        df = read_input_file(str(csv_file))
        assert len(df) == 1
        assert "Название" in df.columns
        assert "Объём SSD;115411" in df.columns

    def test_reads_csv_with_semicolon_separator(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text(
            "Название;Объём SSD\n"
            "Ноутбук 512 ГБ;512 ГБ\n",
            encoding="utf-8"
        )
        df = read_input_file(str(csv_file))
        assert len(df) == 1
        assert "Название" in df.columns

    def test_reads_excel(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        df_orig = pd.DataFrame({
            "Название": ["Ноутбук"],
            "Объём SSD;115411": ["512 ГБ"],
        })
        df_orig.to_excel(str(xlsx_file), index=False)
        df = read_input_file(str(xlsx_file))
        assert len(df) == 1
        assert "Название" in df.columns

    def test_raises_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            read_input_file(str(tmp_path / "nonexistent.csv"))

    def test_raises_value_error_for_unsupported_format(self, tmp_path):
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("test")
        with pytest.raises(ValueError, match="Непідтримуваний формат"):
            read_input_file(str(txt_file))

    def test_all_values_read_as_strings(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("Число\n42\n3.14\n", encoding="utf-8")
        df = read_input_file(str(csv_file))
        # pandas 3.x повертає StringDtype; перевіряємо, що значення — рядки
        assert pd.api.types.is_string_dtype(df["Число"])


# ===========================================================================
# Тести Модуля 8: validate_feed (інтеграційні)
# ===========================================================================


class TestValidateFeed:
    def make_csv(self, tmp_path: Path, rows: list[dict]) -> str:
        """Допоміжний метод: створює тимчасовий CSV-файл."""
        df = pd.DataFrame(rows)
        csv_path = str(tmp_path / "input.csv")
        df.to_csv(csv_path, index=False, encoding="utf-8")
        return csv_path

    def test_creates_output_file(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {"Название": "Ноутбук 512 ГБ SSD", "Объём SSD;115411": "512 ГБ"},
        ])
        out = str(tmp_path / "result.xlsx")
        validate_feed(csv_path, out)
        assert Path(out).exists()

    def test_no_errors_for_matching_values(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {"Название": "Ноутбук 512 ГБ SSD", "Объём SSD;115411": "512 ГБ"},
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert "Помилки" in df.columns
        assert df["Помилки"].iloc[0] == ""

    def test_detects_ssd_conflict(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {
                "Название": "Ноутбук 128 ГБ SSD",
                "Краткое описание": "Ноутбук із 256 ГБ SSD",
                "Объём SSD;115411": "512 ГБ",
            }
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert df["Помилки"].iloc[0] != ""
        assert "Конфлікт SSD" in df["Помилки"].iloc[0]

    def test_detects_ram_conflict(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {
                "Название": "Ноутбук 4 GB RAM",
                "Объем установленной оперативной памяти;20863": "32 GB",
            }
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert "Конфлікт RAM" in df["Помилки"].iloc[0]

    def test_detects_conflict_in_html_description(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {
                "Название": "Ноутбук ASUS",
                "Описание;1": "<p>Ноутбук із <b>128 ГБ</b> SSD</p>",
                "Объём SSD;115411": "512 ГБ",
            }
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert "Конфлікт SSD" in df["Помилки"].iloc[0]

    def test_multiple_conflicts_joined_by_pipe(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {
                "Название": "Ноутбук 128 ГБ SSD 4 GB RAM",
                "Объём SSD;115411": "512 ГБ",
                "Объем установленной оперативной памяти;20863": "32 GB",
            }
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        errors = df["Помилки"].iloc[0]
        assert "Конфлікт SSD" in errors
        assert "Конфлікт RAM" in errors
        assert " | " in errors

    def test_no_conflict_when_text_has_no_memory(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {
                "Название": "Ноутбук ASUS VivoBook",
                "Краткое описание": "Потужний тонкий ноутбук",
                "Объём SSD;115411": "512 ГБ",
            }
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert df["Помилки"].iloc[0] == ""

    def test_multiple_rows_only_conflicting_marked(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            # Рядок 0: без конфлікту
            {
                "Название": "Ноутбук 512 ГБ SSD",
                "Объём SSD;115411": "512 ГБ",
            },
            # Рядок 1: з конфліктом
            {
                "Название": "Ноутбук 128 ГБ SSD",
                "Объём SSD;115411": "512 ГБ",
            },
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert df["Помилки"].iloc[0] == ""
        assert df["Помилки"].iloc[1] != ""

    def test_output_xlsx_has_pomylky_column(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {"Название": "Ноутбук", "Объём SSD;115411": "512 ГБ"},
        ])
        out = str(tmp_path / "result.xlsx")
        validate_feed(csv_path, out)
        df_out = pd.read_excel(out)
        assert "Помилки" in df_out.columns

    def test_sample_data_csv(self):
        """Перевірка на реальному файлі зразкових даних."""
        sample = Path(__file__).parent.parent / "sample_data.csv"
        if not sample.exists():
            pytest.skip("sample_data.csv не знайдено")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            df = validate_feed(str(sample), out)
            # SKU002: назва каже 128ГБ, опис — 256ГБ, характеристика — 128ГБ → конфлікт в описі
            # SKU003: назва і характеристика 1ТБ, опис — 512ГБ → конфлікт
            # SKU004: назва 4ГБ RAM, опис і характеристика RAM різні → конфлікт
            errors_count = (df["Помилки"] != "").sum()
            assert errors_count > 0, "Зразкові дані повинні містити конфлікти"
        finally:
            os.unlink(out)

    def test_detects_diagonal_conflict(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {
                "Название": '14" ноутбук',
                "Діагональ екрана": "15.6",
            }
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert "Конфлікт Діагональ екрана" in df["Помилки"].iloc[0]

    def test_no_diagonal_conflict_when_matching(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {
                "Название": '15.6" ноутбук',
                "Діагональ екрана": "15.6",
            }
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert df["Помилки"].iloc[0] == ""

    def test_detects_matrix_conflict(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {
                "Название": "Ноутбук з TN матрицею",
                "Тип матриці": "IPS",
            }
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert "Конфлікт Тип матриці" in df["Помилки"].iloc[0]

    def test_no_matrix_conflict_when_matching(self, tmp_path):
        csv_path = self.make_csv(tmp_path, [
            {
                "Название": "Ноутбук IPS дисплей",
                "Тип матриці": "IPS",
            }
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert df["Помилки"].iloc[0] == ""

    def test_rules_skipped_when_column_absent(self, tmp_path):
        # Файл без колонки діагоналі — правило "Діагональ екрана" не застосовується
        csv_path = self.make_csv(tmp_path, [
            {"Название": "Ноутбук", "Объём SSD;115411": "512 ГБ"},
        ])
        out = str(tmp_path / "result.xlsx")
        df = validate_feed(csv_path, out)
        assert df["Помилки"].iloc[0] == ""




class TestCommandLine:
    """Тести для main() та аргументів командного рядка."""

    def test_main_with_optional_output(self, tmp_path):
        """Перевірка, що main() працює з опціональним аргументом output."""
        import subprocess
        import sys

        csv_path = tmp_path / "test_input.csv"
        df = pd.DataFrame([
            {"Название": "Ноутбук 512 ГБ SSD", "Объём SSD;115411": "512 ГБ"},
        ])
        df.to_csv(csv_path, index=False, encoding="utf-8")

        # Запускаємо без вказання вихідного файлу
        result = subprocess.run(
            [sys.executable, "checkr.py", str(csv_path)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent,
        )

        assert result.returncode == 0, f"Помилка: {result.stderr}"

        # Перевіряємо, що створився файл з автоматичною назвою
        # Файл має бути створений у тій же директорії, що й вхідний файл
        expected_output = tmp_path / "test_input_result.xlsx"
        assert expected_output.exists(), f"Вихідний файл не створено: {expected_output}"

        try:
            # Перевіряємо вміст
            df_out = pd.read_excel(expected_output)
            assert "Помилки" in df_out.columns
        finally:
            # Видаляємо створений файл після тесту
            if expected_output.exists():
                expected_output.unlink()


# ===========================================================================
# Тести розширеної числової перевірки
# ===========================================================================


class TestMemoryToMb:
    """Тести функції memory_to_mb: конвертація до МБ."""

    def test_mb_returns_exact(self):
        assert memory_to_mb("256МБ") == 256.0

    def test_gb_multiplied_by_1024(self):
        assert memory_to_mb("512ГБ") == 512.0 * 1024

    def test_tb_multiplied_by_1024_squared(self):
        assert memory_to_mb("1ТБ") == 1024.0 * 1024.0

    def test_fractional_gb(self):
        assert memory_to_mb("0.5ГБ") == 0.5 * 1024

    def test_unknown_unit_returns_none(self):
        assert memory_to_mb("512ХБ") is None

    def test_tb_equals_1024_gb(self):
        """1ТБ і 1024ГБ мають давати однаковий результат."""
        assert memory_to_mb("1ТБ") == memory_to_mb("1024ГБ")

    def test_025_tb_equals_256_gb(self):
        """0.25ТБ == 256ГБ (бінарне переведення)."""
        assert memory_to_mb("0.25ТБ") == memory_to_mb("256ГБ")


class TestUnitAwareMemoryComparison:
    """Тести: check_conflicts не видає конфлікт при еквівалентних значеннях у різних одиницях."""

    def _ssd_rule(self) -> dict:
        return {
            "label": "SSD",
            "checker_type": "memory",
            "char_hints": ["Об'єм SSD"],
            "text_hints": ["Название"],
        }

    def test_tb_in_name_gb_in_char_no_conflict(self):
        """«1ТБ» у назві та «1024ГБ» у характеристиці — не конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук 1 ТБ SSD",
            "Об'єм SSD": "1024 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error == ""

    def test_025_tb_in_name_256_gb_in_char_no_conflict(self):
        """«0.25ТБ» у назві та «256ГБ» у характеристиці — не конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук 0.25 ТБ SSD",
            "Об'єм SSD": "256 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error == ""

    def test_different_values_still_conflict(self):
        """Різні значення (512ГБ vs 256ГБ) — конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук 512 ГБ SSD",
            "Об'єм SSD": "256 ГБ",
        })
        error, cols = check_conflicts(row, list(row.index), self._ssd_rule())
        assert error != ""
        assert "Конфлікт SSD" in error


class TestTypoDetection:
    """Тести виявлення типових опечаток (зайвий нуль)."""

    def _battery_rule(self) -> dict:
        return {
            "label": "Ємність акумулятора",
            "checker_type": "battery",
            "char_hints": ["Акумулятор"],
            "text_hints": ["Название"],
        }

    def test_extra_zero_typo_hint_in_error(self):
        """5120 мАг у назві vs 512 мАг у характеристиці — підказка про опечатку."""
        row = pd.Series({
            "Название": "Планшет 5120 mAh",
            "Акумулятор": "512 mAh",
        })
        error, cols = check_conflicts(row, list(row.index), self._battery_rule())
        assert error != ""
        assert "можливо зайвий нуль" in error

    def test_no_typo_hint_for_normal_conflict(self):
        """5000 мАг vs 6000 мАг — звичайний конфлікт без підказки про опечатку."""
        row = pd.Series({
            "Название": "Планшет 5000 mAh",
            "Акумулятор": "6000 mAh",
        })
        error, cols = check_conflicts(row, list(row.index), self._battery_rule())
        assert error != ""
        assert "можливо зайвий нуль" not in error


# ===========================================================================
# Тести ємності акумулятора (battery)
# ===========================================================================


class TestNormalizeBatteryValue:
    """Тести нормалізації значень ємності акумулятора."""

    def test_mah_latin(self):
        assert normalize_battery_value("5000 mAh") == "5000мАг"

    def test_mah_cyrillic(self):
        assert normalize_battery_value("5000мАг") == "5000мАг"

    def test_mah_dot_notation(self):
        assert normalize_battery_value("5000мА·год") == "5000мАг"

    def test_ah_latin(self):
        assert normalize_battery_value("5Ah") == "5Аг"

    def test_ah_cyrillic(self):
        assert normalize_battery_value("5Аг") == "5Аг"

    def test_ah_dot_notation(self):
        assert normalize_battery_value("5А·год") == "5Аг"

    def test_fractional_ah(self):
        assert normalize_battery_value("5.5 Ah") == "5.5Аг"

    def test_unknown_returns_uppercase(self):
        assert normalize_battery_value("unknown") == "UNKNOWN"

    def test_trailing_zeros_stripped(self):
        assert normalize_battery_value("5000.0 mAh") == "5000мАг"


class TestExtractBatteryMatches:
    """Тести пошуку значень ємності акумулятора у тексті."""

    def test_finds_mah_value(self):
        result = extract_battery_matches("Планшет із акумулятором 5000 mAh")
        assert len(result) == 1
        assert result[0][1] == "5000мАг"

    def test_finds_ah_value(self):
        result = extract_battery_matches("Акумулятор 5 Ah")
        assert len(result) == 1
        assert result[0][1] == "5Аг"

    def test_finds_cyrillic_mah(self):
        result = extract_battery_matches("Зарядка 5000 мАг")
        assert len(result) == 1
        assert result[0][1] == "5000мАг"

    def test_empty_string_returns_empty(self):
        assert extract_battery_matches("") == []

    def test_none_returns_empty(self):
        assert extract_battery_matches(None) == []

    def test_no_battery_value_returns_empty(self):
        assert extract_battery_matches("Ноутбук 512 ГБ SSD") == []

    def test_multiple_values(self):
        result = extract_battery_matches("5000 mAh або 5 Ah")
        norms = [n for _, n in result]
        assert "5000мАг" in norms
        assert "5Аг" in norms


class TestBatteryToMah:
    """Тести конвертації значень акумулятора до мАг."""

    def test_mah_unit_returns_same(self):
        assert battery_to_mah("5000мАг") == 5000.0

    def test_ah_unit_multiplied_by_1000(self):
        assert battery_to_mah("5Аг") == 5000.0

    def test_5ah_equals_5000mah(self):
        assert battery_to_mah("5Аг") == battery_to_mah("5000мАг")

    def test_unknown_unit_returns_none(self):
        assert battery_to_mah("5000Вт") is None

    def test_fractional_ah(self):
        assert battery_to_mah("5.5Аг") == 5500.0


class TestBatteryUnitAwareComparison:
    """Тести: check_conflicts не видає конфлікт при еквівалентних значеннях mAh/Ah."""

    def _battery_rule(self) -> dict:
        return {
            "label": "Ємність акумулятора",
            "checker_type": "battery",
            "char_hints": ["Акумулятор"],
            "text_hints": ["Название"],
        }

    def test_5ah_in_name_5000mah_in_char_no_conflict(self):
        """«5 Ah» у назві та «5000 mAh» у характеристиці — не конфлікт."""
        row = pd.Series({
            "Название": "Телефон 5 Ah акумулятор",
            "Акумулятор": "5000 mAh",
        })
        error, cols = check_conflicts(row, list(row.index), self._battery_rule())
        assert error == ""

    def test_5000mah_in_name_5ah_in_char_no_conflict(self):
        """«5000 mAh» у назві та «5 Ah» у характеристиці — не конфлікт."""
        row = pd.Series({
            "Название": "Телефон 5000 mAh акумулятор",
            "Акумулятор": "5 Ah",
        })
        error, cols = check_conflicts(row, list(row.index), self._battery_rule())
        assert error == ""

    def test_different_battery_values_conflict(self):
        """5000 mAh vs 6000 mAh — конфлікт."""
        row = pd.Series({
            "Название": "Телефон 5000 mAh акумулятор",
            "Акумулятор": "6000 mAh",
        })
        error, cols = check_conflicts(row, list(row.index), self._battery_rule())
        assert error != ""
        assert "Конфлікт Ємність акумулятора" in error

    def test_battery_rule_in_validate_feed(self, tmp_path):
        """Інтеграційний тест: виявлення конфлікту ємності акумулятора."""
        df = pd.DataFrame([{
            "Название": "Телефон 3000 mAh акумулятор",
            "Ємність акумулятора": "5000 mAh",
        }])
        csv_path = str(tmp_path / "input.csv")
        df.to_csv(csv_path, index=False, encoding="utf-8")
        out = str(tmp_path / "result.xlsx")
        result_df = validate_feed(csv_path, out)
        assert "Конфлікт Ємність акумулятора" in result_df["Помилки"].iloc[0]


# ===========================================================================
# Тести розширеної перевірки ваги (кг/г)
# ===========================================================================


class TestWeightToGrams:
    """Тести конвертації ваги до грамів."""

    def test_kg_multiplied_by_1000(self):
        assert weight_to_grams("1.5кг") == 1500.0

    def test_grams_returns_same(self):
        assert weight_to_grams("1500г") == 1500.0

    def test_1500g_equals_1_5kg(self):
        assert weight_to_grams("1500г") == weight_to_grams("1.5кг")

    def test_unknown_unit_returns_none(self):
        assert weight_to_grams("1.5фунт") is None

    def test_2kg_returns_2000g(self):
        assert weight_to_grams("2кг") == 2000.0


class TestExtractWeightMatchesWithGrams:
    """Тести пошуку значень ваги з підтримкою грамів."""

    def test_finds_grams(self):
        result = extract_weight_matches("Вага 1500г")
        assert len(result) == 1
        assert result[0][1] == "1500г"

    def test_finds_gr(self):
        result = extract_weight_matches("Вага 1500гр")
        assert len(result) == 1
        assert result[0][1] == "1500г"

    def test_finds_kg_still_works(self):
        result = extract_weight_matches("Вага 1.5 кг")
        assert len(result) == 1
        assert result[0][1] == "1.5кг"

    def test_does_not_match_gb(self):
        """«ГБ» не повинно розпізнаватися як грами."""
        result = extract_weight_matches("512 ГБ SSD")
        assert result == []

    def test_does_not_match_gb_latin(self):
        """«GB» не повинно розпізнаватися як грами."""
        result = extract_weight_matches("256 GB RAM")
        assert result == []


class TestWeightUnitAwareComparison:
    """Тести: check_conflicts не видає конфлікт при еквівалентних кг/г."""

    def _weight_rule(self) -> dict:
        return {
            "label": "Вага",
            "checker_type": "weight",
            "char_hints": ["Вага"],
            "text_hints": ["Название"],
        }

    def test_1500g_in_name_1_5kg_in_char_no_conflict(self):
        """«1500г» у назві та «1.5кг» у характеристиці — не конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук 1500г",
            "Вага": "1.5кг",
        })
        error, cols = check_conflicts(row, list(row.index), self._weight_rule())
        assert error == ""

    def test_1_5kg_in_name_1500g_in_char_no_conflict(self):
        """«1.5кг» у назві та «1500г» у характеристиці — не конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук 1.5 кг",
            "Вага": "1500г",
        })
        error, cols = check_conflicts(row, list(row.index), self._weight_rule())
        assert error == ""

    def test_different_weight_values_conflict(self):
        """1.5кг vs 2.0кг — конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук 1.5 кг",
            "Вага": "2.0кг",
        })
        error, cols = check_conflicts(row, list(row.index), self._weight_rule())
        assert error != ""
        assert "Конфлікт Вага" in error


# ===========================================================================
# Тести семантичної перевірки
# ===========================================================================


class TestCheckSemanticConflicts:
    """Тести функції check_semantic_conflicts."""

    def test_wireless_wired_contradiction_detected(self):
        """«бездротовий» у назві та «з дротом» в описі — конфлікт."""
        row = pd.Series({
            "Название": "Бездротовий ноутбук",
            "Краткое описание": "Пристрій з дротом",
        })
        error, cols = check_semantic_conflicts(row, list(row.index))
        assert error != ""
        assert "Бездротовий/Дротовий" in error
        assert "Название" in cols
        assert "Краткое описание" in cols

    def test_wireless_wired_ru_detected(self):
        """«беспроводной» у назві та «проводной» в описі — конфлікт."""
        row = pd.Series({
            "Название": "Беспроводной адаптер",
            "Краткое описание": "Проводной интерфейс",
        })
        error, cols = check_semantic_conflicts(row, list(row.index))
        assert error != ""
        assert "Бездротовий/Дротовий" in error

    def test_no_contradiction_for_wireless_only(self):
        """Лише «бездротовий» без протилежного — не конфлікт."""
        row = pd.Series({
            "Название": "Бездротовий ноутбук",
            "Краткое описание": "Підключення через wi-fi",
        })
        error, cols = check_semantic_conflicts(row, list(row.index))
        assert "Бездротовий/Дротовий" not in error

    def test_no_contradiction_for_wired_only(self):
        """Лише «з дротом» без протилежного — не конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук з дротом",
            "Краткое описание": "Провідне з'єднання",
        })
        error, cols = check_semantic_conflicts(row, list(row.index))
        assert "Бездротовий/Дротовий" not in error

    def test_core_count_conflict_detected(self):
        """«12-ядерний» у назві та «4-ядерний» в описі — конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук 12-ядерний процесор",
            "Краткое описание": "Потужний 4-ядерний CPU",
        })
        error, cols = check_semantic_conflicts(row, list(row.index))
        assert error != ""
        assert "Конфлікт кількості ядер" in error

    def test_core_count_same_no_conflict(self):
        """Однакова кількість ядер у всіх полях — не конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук 8-ядерний процесор",
            "Краткое описание": "Потужний 8-ядерний CPU",
        })
        error, cols = check_semantic_conflicts(row, list(row.index))
        assert "Конфлікт кількості ядер" not in error

    def test_core_count_single_field_no_conflict(self):
        """Кількість ядер лише в одному полі — не конфлікт."""
        row = pd.Series({
            "Название": "Ноутбук 8-ядерний процесор",
            "Краткое описание": "Потужний ноутбук для роботи",
        })
        error, cols = check_semantic_conflicts(row, list(row.index))
        assert "Конфлікт кількості ядер" not in error

    def test_empty_row_no_conflict(self):
        """Порожній рядок — немає конфліктів."""
        row = pd.Series({"Название": "", "Краткое описание": ""})
        error, cols = check_semantic_conflicts(row, list(row.index))
        assert error == ""

    def test_semantic_error_in_errors_column(self, tmp_path):
        """Інтеграційний тест: семантичний конфлікт записується в колонку Помилки."""
        df = pd.DataFrame([{
            "Название": "Бездротовий ноутбук",
            "Краткое описание": "Пристрій з дротом",
        }])
        csv_path = str(tmp_path / "input.csv")
        df.to_csv(csv_path, index=False, encoding="utf-8")
        out = str(tmp_path / "result.xlsx")
        result_df = validate_feed(csv_path, out)
        assert "Семантична перевірка" in result_df["Помилки"].iloc[0]
        assert "Бездротовий/Дротовий" in result_df["Помилки"].iloc[0]

    def test_semantic_conflict_highlighted_in_excel(self, tmp_path):
        """Інтеграційний тест: конфліктні колонки підсвічені в Excel."""
        from openpyxl import load_workbook
        df = pd.DataFrame([{
            "Название": "Бездротовий ноутбук",
            "Краткое описание": "Пристрій з дротом",
        }])
        csv_path = str(tmp_path / "input.csv")
        df.to_csv(csv_path, index=False, encoding="utf-8")
        out = str(tmp_path / "result.xlsx")
        validate_feed(csv_path, out)
        wb = load_workbook(out)
        ws = wb.active
        # Переконуємося, що клітинка Помилки підсвічена (жовтий)
        header = {cell.value: cell.column for cell in ws[1]}
        pomylky_col = header.get("Помилки")
        assert pomylky_col is not None
        cell = ws.cell(row=2, column=pomylky_col)
        assert cell.fill.fgColor.rgb != "00000000"

    def test_cores_english_detected(self):
        """«12 cores» у назві та «4 cores» в описі — конфлікт."""
        row = pd.Series({
            "Название": "Laptop with 12 cores CPU",
            "Краткое описание": "4 cores processor",
        })
        error, cols = check_semantic_conflicts(row, list(row.index))
        assert "Конфлікт кількості ядер" in error


class TestValidationRulesBattery:
    """Тести правил валідації: ємність акумулятора."""

    def test_has_battery_rule(self):
        labels = [r["label"] for r in VALIDATION_RULES]
        assert "Ємність акумулятора" in labels

    def test_battery_rule_has_correct_checker_type(self):
        battery_rule = next(
            r for r in VALIDATION_RULES if r["label"] == "Ємність акумулятора"
        )
        assert battery_rule["checker_type"] == "battery"

    def test_battery_rule_has_mah_hint(self):
        battery_rule = next(
            r for r in VALIDATION_RULES if r["label"] == "Ємність акумулятора"
        )
        assert any("mAh" in hint or "акумулятор" in hint.lower()
                   for hint in battery_rule["char_hints"])

