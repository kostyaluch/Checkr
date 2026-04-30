"""checkr.py — Автоматична валідація товарного фіду e-commerce.

Програма зчитує файл (CSV або Excel) зі списком товарів і знаходить
логічні розбіжності між назвою, описами та характеристиками одного й
того ж товару (наприклад, конфлікти об'єму SSD, оперативної пам'яті,
діагоналі екрана, ваги, роздільної здатності, типу матриці тощо).

Правила валідації зберігаються у validation_rules.py — їх легко
розширювати, не змінюючи основну логіку програми.

Використання з командного рядка:
    python checkr.py products.csv result.xlsx
    python checkr.py products.xlsx result.xlsx
"""

import argparse
import itertools
import math
import re
import sys
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from validation_rules import VALIDATION_RULES
from language_detector import check_language_consistency

# ---------------------------------------------------------------------------
# Константи: параметри контекстно-залежного пошуку пам'яті
# ---------------------------------------------------------------------------

# Бонус відстані для ключових слів, що йдуть ПІСЛЯ значення пам'яті.
# Наприклад, "16 GB RAM" — слово "RAM" йде після значення.
# Цей бонус дає пріоритет таким випадкам перед "SSD ... 16 GB".
KEYWORD_AFTER_VALUE_BONUS = 5

# Максимальна відстань (у символах) між значенням пам'яті та ключовим словом.
# Якщо ключове слово далі — воно не враховується.
MAX_CONTEXT_DISTANCE = 50

# ---------------------------------------------------------------------------
# Константи: кольори для підсвітки у вихідному Excel-файлі
# ---------------------------------------------------------------------------

# Червоний (світлий) — для клітинок із конфліктними значеннями
CONFLICT_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

# Жовтий — для клітинки колонки "Помилки"
ERROR_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


# ===========================================================================
# Модуль 1: Очищення тексту від HTML
# ===========================================================================


def clean_html(text: str) -> str:
    """Видаляє HTML-теги з рядка тексту.

    Використовує BeautifulSoup для надійного парсингу HTML.
    Якщо вхід не є рядком (None, число тощо) — повертає порожній рядок.

    Аргументи:
        text: Рядок, що може містити HTML-теги.

    Повертає:
        Рядок без HTML-тегів із збереженим текстовим вмістом.

    Приклад:
        clean_html("<p>Hello <b>World</b></p>")  →  "Hello World"
        clean_html(None)                          →  ""
    """
    if not isinstance(text, str):
        return ""
    if not text.strip():
        return text
    try:
        # lxml — швидший парсер, якщо встановлений
        return BeautifulSoup(text, "lxml").get_text(separator=" ").strip()
    except Exception:
        # Резервний варіант — вбудований парсер Python
        return BeautifulSoup(text, "html.parser").get_text(separator=" ").strip()


# ===========================================================================
# Модуль 2: Пошук значень пам'яті за допомогою регулярних виразів
# ===========================================================================

# Регулярний вираз для пошуку значень об'єму пам'яті.
# Підтримує:
#   Кириличні одиниці: ГБ (гігабайти), ТБ (терабайти), МБ (мегабайти)
#   Латинські одиниці: GB, TB, MB, GiB, TiB, MiB
#   Дробові значення:  1.5 ТБ, 2,5 GB
#   З пробілом та без: "512ГБ", "512 GB"
#
# Порядок альтернатив у групі одиниць важливий: більш специфічні — першими
# (TiB перед TB, GiB перед GB, MiB перед MB), щоб не захопити лише першу літеру.
_MEMORY_RE = re.compile(
    r"(?<!\d)"                              # не передує цифра (не частина більшого числа)
    r"(\d+(?:[.,]\d+)?)"                    # числова частина: ціле або дробове
    r"\s*"                                  # необов'язковий пробіл між числом та одиницею
    r"(ТБ|TiB|TB|ГБ|GiB|GB|МБ|MiB|MB)"    # одиниці виміру (порядок важливий!)
    r"(?!\w)",                              # не слідує символ слова (немає "GBps")
    re.IGNORECASE,
)

# Маппінг: нижній регістр одиниці → стандартне кириличне позначення.
# Щоб додати нову одиницю (наприклад, "ГБ/s"), додайте запис тут.
_UNIT_NORM: dict[str, str] = {
    "тб": "ТБ", "tb": "ТБ", "tib": "ТБ",
    "гб": "ГБ", "gb": "ГБ", "gib": "ГБ",
    "мб": "МБ", "mb": "МБ", "mib": "МБ",
}

# Мультиплікатори для перетворення одиниць пам'яті до МБ (мегабайт).
# Використовується для порівняння значень у різних одиницях (наприклад, 1ТБ = 1024ГБ).
_MEMORY_UNIT_TO_MB: dict[str, float] = {
    "МБ": 1.0,
    "ГБ": 1024.0,
    "ТБ": 1024.0 * 1024.0,
}


def normalize_memory_value(raw: str) -> str:
    """Нормалізує одне значення пам'яті до стандартного формату «<число><одиниця>».

    Перетворює різні варіанти написання до єдиного стандарту:
      - Замінює латинські позначення на кириличні (GB → ГБ, TB → ТБ).
      - Прибирає пробіл між числом та одиницею.
      - Нормалізує роздільник дробу: кому замінює на крапку.
      - Прибирає зайві нулі після коми: 1.0 → 1, 1.50 → 1.5.
    Якщо рядок не відповідає шаблону — повертає оригінал у верхньому регістрі.

    Аргументи:
        raw: Рядок виду "512 GB", "1.5ТБ", "256мб", "2,5 TiB" тощо.

    Повертає:
        Нормалізований рядок, наприклад: "512ГБ", "1.5ТБ", "256МБ".

    Приклад:
        normalize_memory_value("512 GB")   →  "512ГБ"
        normalize_memory_value("1.5 TB")   →  "1.5ТБ"
        normalize_memory_value("2,5 ГБ")   →  "2.5ГБ"
        normalize_memory_value("unknown")  →  "UNKNOWN"
    """
    m = _MEMORY_RE.search(raw)
    if not m:
        return raw.strip().upper()

    # Нормалізуємо числову частину
    number = m.group(1).replace(",", ".")
    if "." in number:
        number = number.rstrip("0").rstrip(".")

    # Нормалізуємо одиницю: шукаємо в маппінгу, або використовуємо верхній регістр
    unit = _UNIT_NORM.get(m.group(2).lower(), m.group(2).upper())
    return f"{number}{unit}"


def memory_to_mb(normalized: str) -> float | None:
    """Перетворює нормалізоване значення пам'яті (наприклад, «256ГБ») до МБ.

    Підтримує одиниці: МБ, ГБ, ТБ (після нормалізації через normalize_memory_value).

    Аргументи:
        normalized: Нормалізований рядок виду «512ГБ», «1.5ТБ», «256МБ».

    Повертає:
        Числове значення в МБ або None, якщо рядок не вдалось розпарсити.

    Приклад:
        memory_to_mb("512ГБ")  →  524288.0
        memory_to_mb("1ТБ")    →  1048576.0
        memory_to_mb("256МБ")  →  256.0
    """
    for unit, multiplier in _MEMORY_UNIT_TO_MB.items():
        if normalized.endswith(unit):
            try:
                return float(normalized[: -len(unit)]) * multiplier
            except ValueError:
                return None
    return None


def extract_memory_matches(text: str) -> list[tuple[str, str]]:
    """Знаходить усі значення об'єму пам'яті у тексті.

    Повертає список пар (оригінал, нормалізоване), де:
      - оригінал      — рядок так, як знайдений у тексті (напр. "512 GB")
      - нормалізоване — стандартне представлення (напр. "512ГБ")

    Аргументи:
        text: Рядок, у якому шукаємо значення.

    Повертає:
        Список пар (str, str). Порожній список, якщо нічого не знайдено.

    Приклад:
        extract_memory_matches("Ноутбук 512 ГБ SSD та 32 GB RAM")
        →  [("512 ГБ", "512ГБ"), ("32 GB", "32ГБ")]
    """
    if not isinstance(text, str) or not text.strip():
        return []
    return [
        (f"{num}{unit}", normalize_memory_value(f"{num} {unit}"))
        for num, unit in _MEMORY_RE.findall(text)
    ]


def extract_memory_values(text: str) -> list[str]:
    """Зручна обгортка над extract_memory_matches: повертає лише нормалізовані значення.

    Аргументи:
        text: Рядок, у якому шукаємо значення.

    Повертає:
        Список нормалізованих значень, наприклад ["512ГБ", "32ГБ"].

    Приклад:
        extract_memory_values("Ноутбук 512 ГБ SSD та 32 GB RAM")  →  ["512ГБ", "32ГБ"]
        extract_memory_values("Просто текст")                      →  []
    """
    return [norm for _, norm in extract_memory_matches(text)]


def extract_memory_matches_with_context(
    text: str, context_keywords: list[str] | None = None
) -> list[tuple[str, str]]:
    """Знаходить значення об'єму пам'яті у тексті з урахуванням контексту.

    Якщо вказані ключові слова (context_keywords), повертає лише ті значення пам'яті,
    які знаходяться поруч із цими словами. Це дозволяє розрізняти RAM, SSD, VRAM тощо.

    Алгоритм:
      1. Знаходить усі значення пам'яті у тексті.
      2. Для кожного значення перевіряє, чи є хоча б одне ключове слово в радіусі
         MAX_CONTEXT_DISTANCE символів (після коригування відстані).
      3. Додатково, якщо є конкуруючі ключові слова (SSD, RAM тощо) ближче ніж цільове,
         значення не додається (щоб уникнути плутанини).
      4. Якщо context_keywords не вказано або порожній — повертає всі знайдені значення.

    Аргументи:
        text:             Рядок, у якому шукаємо значення.
        context_keywords: Список ключових слів для фільтрації (наприклад, ["SSD", "накопичувач"]).
                          Пошук нечутливий до регістру.

    Повертає:
        Список пар (оригінал, нормалізоване). Порожній список, якщо нічого не знайдено.

    Приклад:
        extract_memory_matches_with_context(
            "Ноутбук із 512 ГБ SSD та 16 GB RAM",
            ["SSD", "накопичувач"]
        )
        →  [("512 ГБ", "512ГБ")]  # Лише SSD, RAM ігнорується

        extract_memory_matches_with_context(
            "Ноутбук із 512 ГБ SSD та 16 GB RAM",
            ["RAM", "оперативна", "оперативной"]
        )
        →  [("16 GB", "16ГБ")]  # Лише RAM, SSD ігнорується

        extract_memory_matches_with_context("Ноутбук 512 ГБ SSD", None)
        →  [("512 ГБ", "512ГБ")]  # Без фільтрації
    """
    if not isinstance(text, str) or not text.strip():
        return []

    # Якщо контекст не вказано — повертаємо всі значення
    if not context_keywords:
        return extract_memory_matches(text)

    # Нормалізуємо ключові слова до нижнього регістру
    keywords_lower = [kw.lower() for kw in context_keywords]

    # Список конкуруючих ключових слів (інші СПЕЦИФІЧНІ типи пам'яті, які НЕ є цільовими)
    # Використовується для виключення значень, якщо інший тип пам'яті ближче
    # НЕ включаємо загальні слова на кшталт "memory", "памят", тільки специфічні типи
    competing_keywords = [
        "ssd", "ссд", "накопичувач", "накопитель", "твердотіл", "твердотел",
        "ram", "озу", "оперативн",
        "vram", "відеопам", "видеопам",
        "hdd", "жорстк", "жестк",
    ]
    
    # Видаляємо з конкурентів ті, що є в цільових keywords
    competitors = [kw for kw in competing_keywords if kw not in keywords_lower]
    
    text_lower = text.lower()
    results: list[tuple[str, str]] = []

    # Знаходимо всі збіги пам'яті
    for match in _MEMORY_RE.finditer(text):
        num, unit = match.groups()
        value_start = match.start()
        value_end = match.end()

        # Знаходимо найближче цільове ключове слово
        # Пріоритет: спочатку шукаємо ПІСЛЯ значення (в межах 15 символів),
        # потім ДО значення
        min_target_distance = math.inf
        for keyword in keywords_lower:
            idx = 0
            while True:
                idx = text_lower.find(keyword, idx)
                if idx == -1:
                    break
                
                # Відстань залежить від того, де знаходиться ключове слово
                if idx >= value_end:
                    # Ключове слово ПІСЛЯ значення - це найбільш природно
                    distance = idx - value_end
                    # Застосовуємо бонус для слів після значення
                    # щоб "4 GB RAM" виграло проти "SSD ... 4 GB"
                    distance = max(0, distance - KEYWORD_AFTER_VALUE_BONUS)
                else:
                    # Ключове слово ДО значення
                    distance = value_start - (idx + len(keyword))
                
                if distance >= 0 and distance < min_target_distance:
                    min_target_distance = distance
                idx += 1

        # Якщо цільове ключове слово далі MAX_CONTEXT_DISTANCE — пропускаємо
        if min_target_distance > MAX_CONTEXT_DISTANCE:
            continue

        # Перевіряємо, чи немає конкуруючого ключового слова ближче
        min_competitor_distance = math.inf
        for keyword in competitors:
            idx = 0
            while True:
                idx = text_lower.find(keyword, idx)
                if idx == -1:
                    break
                
                # Така ж логіка для конкурентів
                if idx >= value_end:
                    distance = idx - value_end
                    distance = max(0, distance - KEYWORD_AFTER_VALUE_BONUS)
                else:
                    distance = value_start - (idx + len(keyword))
                
                if distance >= 0 and distance < min_competitor_distance:
                    min_competitor_distance = distance
                idx += 1

        # Якщо конкурент ближче — пропускаємо це значення
        if min_competitor_distance < min_target_distance:
            continue

        # Всі перевірки пройдено — додаємо значення
        original = f"{num}{unit}"
        normalized = normalize_memory_value(f"{num} {unit}")
        results.append((original, normalized))

    return results


# ===========================================================================
# Модуль 3: Пошук колонок у DataFrame
# ===========================================================================


def find_column(columns: list[str], hint: str) -> str | None:
    """Знаходить назву колонки за підрядком hint.

    Алгоритм пошуку (від найточнішого до найменш точного):
      1. Точний збіг повної назви колонки (нечутливий до регістру).
      2. Точний збіг базової назви — частини до крапки з комою
         (напр. знаходить "Описание;1" за точним збігом base="Описание").
      3. Базова назва починається з підказки з межею слова (\b):
         знаходить "Описание;1" за hint="Описание", але НЕ знаходить
         "Краткое описание", а також знаходить "Назва (ua)" за hint="Назва",
         але НЕ знаходить "Название".
      4. Підказка міститься у базовій назві (часткове входження, fallback):
         знаходить "Объём SSD;115411" за hint="SSD".

    Пошук нечутливий до регістру.

    Аргументи:
        columns: Список назв колонок DataFrame.
        hint:    Підрядок (або повна назва) для пошуку.

    Повертає:
        Назву першої знайденої колонки або None, якщо нічого не знайдено.

    Приклад:
        find_column(["Объём SSD;115411", "Название"], "Объём SSD")   →  "Объём SSD;115411"
        find_column(["Краткое описание", "Описание;1"], "Описание")   →  "Описание;1"
        find_column(["Название", "Назва (ua)"], "відеокарта")          →  None
    """
    hint_lower = hint.lower()

    # Крок 1: точний збіг повної назви колонки
    for col in columns:
        if col.lower() == hint_lower:
            return col

    # Крок 2: точний збіг базової назви (частина до ';')
    for col in columns:
        base = col.split(";")[0].strip().lower()
        if base == hint_lower:
            return col

    # Крок 3: базова назва починається з підказки + межа слова (\b).
    # Наприклад, hint="Описание" знаходить "Описание;1" (base="Описание"),
    # але НЕ знаходить "Краткое описание" (base="Краткое описание").
    start_pattern = re.compile(rf"^{re.escape(hint_lower)}\b", re.UNICODE)
    for col in columns:
        base = col.split(";")[0].strip().lower()
        if start_pattern.match(base):
            return col

    # Крок 4: підказка міститься у базовій назві (найширший fallback).
    # Наприклад, hint="SSD" знаходить "Объём SSD;115411" (base="Объём SSD").
    for col in columns:
        base = col.split(";")[0].strip().lower()
        if hint_lower in base:
            return col

    return None


# ===========================================================================
# Модуль 2b: Пошук значень додаткових типів характеристик
# ===========================================================================


# ---------------------------------------------------------------------------
# Діагональ екрана
# ---------------------------------------------------------------------------

# Патерн для значень діагоналі з одиницею виміру:
# Підтримує: 15.6", 15,6", 14 дюймів, 13.3 inch, 15.6 inches, 15.6″
_DIAGONAL_RE = re.compile(
    r"(?<!\d)"
    r"(\d+(?:[.,]\d+)?)"                                    # числова частина
    r"\s*"
    r"(?:\"|″|дюйм(?:ів|а|и)?|-?inch(?:es)?)"              # одиниця виміру
    r"(?!\d)",
    re.IGNORECASE | re.UNICODE,
)


def extract_screen_diagonal_matches(text: str) -> list[tuple[str, str]]:
    """Знаходить усі значення діагоналі екрана у тексті.

    Розпізнає форми: 15.6", 14 дюймів, 13.3-inch тощо.
    Повертає список пар (оригінал, нормалізоване) у форматі «15.6"».

    Аргументи:
        text: Рядок, у якому шукаємо значення.

    Повертає:
        Список пар (str, str). Порожній список, якщо нічого не знайдено.

    Приклад:
        extract_screen_diagonal_matches('Ноутбук 15.6" IPS')
        →  [('15.6"', '15.6"')]
    """
    if not isinstance(text, str) or not text.strip():
        return []
    results = []
    for m in _DIAGONAL_RE.finditer(text):
        original = m.group(0).strip()
        number = m.group(1).replace(",", ".")
        if "." in number:
            number = number.rstrip("0").rstrip(".")
        normalized = f'{number}"'
        results.append((original, normalized))
    return results


# ---------------------------------------------------------------------------
# Вага пристрою
# ---------------------------------------------------------------------------

# Патерн для значень ваги в кілограмах або грамах:
# Підтримує: 1.5 кг, 2,3 кг, 1.8kg, 2.1 KG, 1500 г, 1200гр
# Порядок важливий: «кг» перед «г», щоб «кг» не захоплювалось як «к» + «г».
_WEIGHT_RE = re.compile(
    r"(?<!\d)"
    r"(\d+(?:[.,]\d+)?)"                  # числова частина
    r"\s*"
    r"(кг|kg|гр\.?|г)"                    # одиниці: кг/kg перед г/гр
    r"(?!\w)",
    re.IGNORECASE | re.UNICODE,
)


def extract_weight_matches(text: str) -> list[tuple[str, str]]:
    """Знаходить усі значення ваги у тексті (кілограми або грами).

    Розпізнає форми: 1.5 кг, 2,3 кг, 1.8kg, 1500г, 1200гр тощо.
    Повертає список пар (оригінал, нормалізоване) у форматі «1.5кг» або «1500г».

    Аргументи:
        text: Рядок, у якому шукаємо значення.

    Повертає:
        Список пар (str, str). Порожній список, якщо нічого не знайдено.

    Приклад:
        extract_weight_matches("Легкий ноутбук вагою 1.5 кг")
        →  [('1.5 кг', '1.5кг')]
        extract_weight_matches("Вага: 1500г")
        →  [('1500г', '1500г')]
    """
    if not isinstance(text, str) or not text.strip():
        return []
    results = []
    for m in _WEIGHT_RE.finditer(text):
        original = m.group(0).strip()
        number = m.group(1).replace(",", ".")
        if "." in number:
            number = number.rstrip("0").rstrip(".")
        unit_raw = m.group(2).lower().rstrip(".")
        # Нормалізуємо одиницю: кг/kg → кг, г/гр → г
        if unit_raw in ("кг", "kg"):
            unit_norm = "кг"
        else:
            unit_norm = "г"
        results.append((original, f"{number}{unit_norm}"))
    return results


def weight_to_grams(normalized: str) -> float | None:
    """Перетворює нормалізоване значення ваги до грамів для порівняння.

    Підтримує одиниці: кг, г (після нормалізації через extract_weight_matches).

    Аргументи:
        normalized: Нормалізований рядок виду «1.5кг», «1500г».

    Повертає:
        Числове значення у грамах або None, якщо рядок не вдалось розпарсити.

    Приклад:
        weight_to_grams("1.5кг")  →  1500.0
        weight_to_grams("1500г")  →  1500.0
        weight_to_grams("2кг")    →  2000.0
    """
    if normalized.endswith("кг"):
        try:
            return float(normalized[:-2]) * 1000
        except ValueError:
            return None
    if normalized.endswith("г"):
        try:
            return float(normalized[:-1])
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Ємність акумулятора (mAh / Ah)
# ---------------------------------------------------------------------------

# Патерн для значень ємності акумулятора:
# Підтримує: 5000 мАг, 5000 mAh, 5000мА·год, 5 Аг, 5Ah, 5А·год
# Порядок важливий: більш специфічні одиниці («мА·год») — першими.
_BATTERY_RE = re.compile(
    r"(?<!\d)"
    r"(\d+(?:[.,]\d+)?)"                                     # числова частина
    r"\s*"
    r"(мА·год|мА·г|мАг|mAh|А·год|А·г|Аг|Ah)"               # одиниці (порядок важливий)
    r"(?!\w)",
    re.IGNORECASE | re.UNICODE,
)

# Маппінг: нижній регістр одиниці → нормалізована форма.
_BATTERY_UNIT_NORM: dict[str, str] = {
    "ма·год": "мАг", "ма·г": "мАг", "маг": "мАг", "mah": "мАг",
    "а·год": "Аг",   "а·г": "Аг",   "аг": "Аг",   "ah": "Аг",
}

# Мультиплікатори для перетворення одиниць ємності до мАг (міліампер-годин).
_BATTERY_UNIT_TO_MAH: dict[str, float] = {
    "мАг": 1.0,
    "Аг":  1000.0,
}


def normalize_battery_value(raw: str) -> str:
    """Нормалізує одне значення ємності акумулятора до стандартного формату.

    Перетворює різні варіанти написання до єдиного стандарту:
      - «mAh», «мА·год» → «мАг»
      - «Ah», «А·год»   → «Аг»
    Якщо рядок не відповідає шаблону — повертає оригінал у верхньому регістрі.

    Аргументи:
        raw: Рядок виду «5000 mAh», «5Ah», «5000мА·год» тощо.

    Повертає:
        Нормалізований рядок, наприклад: «5000мАг», «5Аг».

    Приклад:
        normalize_battery_value("5000 mAh")   →  "5000мАг"
        normalize_battery_value("5Ah")         →  "5Аг"
        normalize_battery_value("5000мА·год")  →  "5000мАг"
    """
    m = _BATTERY_RE.search(raw)
    if not m:
        return raw.strip().upper()
    number = m.group(1).replace(",", ".")
    if "." in number:
        number = number.rstrip("0").rstrip(".")
    unit = _BATTERY_UNIT_NORM.get(m.group(2).lower(), m.group(2))
    return f"{number}{unit}"


def extract_battery_matches(text: str) -> list[tuple[str, str]]:
    """Знаходить усі значення ємності акумулятора у тексті.

    Розпізнає форми: 5000 mAh, 5000мА·год, 5Ah тощо.
    Повертає список пар (оригінал, нормалізоване) у форматі «5000мАг» або «5Аг».

    Аргументи:
        text: Рядок, у якому шукаємо значення.

    Повертає:
        Список пар (str, str). Порожній список, якщо нічого не знайдено.

    Приклад:
        extract_battery_matches("Акумулятор 5000 mAh")
        →  [('5000 mAh', '5000мАг')]
    """
    if not isinstance(text, str) or not text.strip():
        return []
    results = []
    for m in _BATTERY_RE.finditer(text):
        original = m.group(0).strip()
        number = m.group(1).replace(",", ".")
        if "." in number:
            number = number.rstrip("0").rstrip(".")
        unit_key = m.group(2)
        unit = _BATTERY_UNIT_NORM.get(unit_key.lower(), unit_key)
        results.append((original, f"{number}{unit}"))
    return results


def battery_to_mah(normalized: str) -> float | None:
    """Перетворює нормалізоване значення ємності до мАг (міліампер-годин).

    Аргументи:
        normalized: Нормалізований рядок виду «5000мАг» або «5Аг».

    Повертає:
        Числове значення у мАг або None, якщо рядок не вдалось розпарсити.

    Приклад:
        battery_to_mah("5000мАг")  →  5000.0
        battery_to_mah("5Аг")      →  5000.0
    """
    for unit, multiplier in _BATTERY_UNIT_TO_MAH.items():
        if normalized.endswith(unit):
            try:
                return float(normalized[: -len(unit)]) * multiplier
            except ValueError:
                return None
    return None


# ---------------------------------------------------------------------------
# Роздільна здатність екрана
# ---------------------------------------------------------------------------

# Патерн для числової роздільної здатності (NxM або N×M):
# Підтримує: 1920x1080, 1920×1080, 3840 x 2160
_RESOLUTION_NUM_RE = re.compile(
    r"(\d{3,4})\s*[xхх×]\s*(\d{3,4})",
    re.IGNORECASE,
)

# Псевдоніми роздільної здатності → нормалізована форма «ШxВ».
# Довші рядки мають бути раніше, щоб "Full HD" знайшлося раніше "HD".
_RESOLUTION_ALIASES: dict[str, str] = {
    "full hd": "1920x1080",
    "full-hd": "1920x1080",
    "fhd": "1920x1080",
    "ultra hd": "3840x2160",
    "ultra-hd": "3840x2160",
    "uhd": "3840x2160",
    "4k": "3840x2160",
    "qhd": "2560x1440",
    "hd+": "1600x900",
    "wxga+": "1600x900",
    "wxga": "1366x768",
    "hd": "1366x768",
}


def extract_resolution_matches(text: str) -> list[tuple[str, str]]:
    """Знаходить усі значення роздільної здатності у тексті.

    Розпізнає числові значення (1920x1080, 3840×2160) та псевдоніми
    (FHD, Full HD, 4K, QHD, HD+ тощо).
    Повертає список пар (оригінал, нормалізоване) у форматі «1920x1080».
    Довші псевдоніми мають пріоритет над коротшими (напр. "Full HD" над "HD").

    Аргументи:
        text: Рядок, у якому шукаємо значення.

    Повертає:
        Список пар (str, str). Порожній список, якщо нічого не знайдено.

    Приклад:
        extract_resolution_matches("Дисплей FHD 1920x1080")
        →  [('FHD', '1920x1080'), ('1920x1080', '1920x1080')]
    """
    if not isinstance(text, str) or not text.strip():
        return []

    results: list[tuple[str, str]] = []
    seen_norms: set[str] = set()
    matched_ranges: list[tuple[int, int]] = []  # Вже зайняті позиції
    text_lower = text.lower()

    def _overlaps(start: int, end: int) -> bool:
        return any(s < end and start < e for s, e in matched_ranges)

    # Перевіряємо псевдоніми (від довших до коротших, щоб "Full HD" знаходилось
    # раніше "HD" і попередні збіги не перекривалися)
    for alias in sorted(_RESOLUTION_ALIASES, key=len, reverse=True):
        # Використовуємо регулярний вираз з межами слова, щоб "HD" не знаходилось у "HDMI"
        pattern = re.compile(r'\b' + re.escape(alias) + r'\b', re.IGNORECASE)
        match = pattern.search(text_lower)
        if match:
            idx = match.start()
            end = match.end()
            if not _overlaps(idx, end):
                norm = _RESOLUTION_ALIASES[alias]
                if norm not in seen_norms:
                    results.append((text[idx: end], norm))
                    seen_norms.add(norm)
                    matched_ranges.append((idx, end))

    # Шукаємо числові роздільні здатності
    for m in _RESOLUTION_NUM_RE.finditer(text):
        if not _overlaps(m.start(), m.end()):
            norm = f"{m.group(1)}x{m.group(2)}"
            if norm not in seen_norms:
                results.append((m.group(0), norm))
                seen_norms.add(norm)
                matched_ranges.append((m.start(), m.end()))

    return results


# ---------------------------------------------------------------------------
# Список допустимих значень (для текстових характеристик: матриця, GPU тощо)
# ---------------------------------------------------------------------------


def extract_value_list_matches(
    text: str, valid_values: list[str]
) -> list[tuple[str, str]]:
    """Знаходить у тексті будь-яке зі списку допустимих значень.

    Пошук нечутливий до регістру. Збіг визначається як входження рядка
    valid_values[i] у текст. Довші значення перевіряються першими.

    Аргументи:
        text:         Рядок, у якому шукаємо значення.
        valid_values: Список рядків для пошуку (наприклад, ["IPS", "TN", "VA"]).

    Повертає:
        Список пар (оригінал_у_тексті, нормалізоване_верхній_регістр).
        Порожній список, якщо нічого не знайдено.

    Приклад:
        extract_value_list_matches("Матриця IPS, тонкі рамки", ["IPS", "TN"])
        →  [('IPS', 'IPS')]
    """
    if not isinstance(text, str) or not text.strip():
        return []

    text_lower = text.lower()
    results: list[tuple[str, str]] = []
    seen: set[str] = set()

    # Довші значення перевіряємо першими (щоб "Mini-LED" знаходилось раніше "LED")
    for val in sorted(valid_values, key=len, reverse=True):
        val_lower = val.lower()
        if val_lower in seen:
            continue
        idx = text_lower.find(val_lower)
        if idx >= 0:
            results.append((text[idx: idx + len(val)], val.upper()))
            seen.add(val_lower)

    return results


# ---------------------------------------------------------------------------
# Модель процесора (CPU)
# ---------------------------------------------------------------------------

# Регулярний вираз для пошуку моделей процесорів Intel та AMD.
# Підтримує:
#   Intel: Core i3/i5/i7/i9, Core Ultra, Pentium, Celeron, Core 5/7
#   AMD: Ryzen 3/5/7/9, Threadripper, Athlon
# Приклади:
#   "Intel Core i7-12700F", "AMD Ryzen 5 7520U", "Intel Core Ultra 5 125H"
_CPU_MODEL_RE = re.compile(
    r"""
    (?:Intel|AMD)\s+                          # Виробник (обов'язково)
    (?:
        # Intel: Core i3/i5/i7/i9 + модель
        Core\s+[iI](?:3|5|7|9)[-\s]*\d+[A-Z]*
        |
        # Intel: Core Ultra + число + модель
        Core\s+Ultra\s+(?:X)?(?:3|5|7|9)\s+\d+[A-Z]*
        |
        # Intel: Core 5/7 + модель (нові процесори)
        Core\s+(?:3|5|7|9)\s+(?:processor\s+)?\d+[A-Z]*
        |
        # Intel: Pentium, Celeron + модель
        (?:Pentium|Celeron)(?:\s+(?:Gold|Silver))?\s+[A-Z]?\d+[A-Z]*
        |
        # AMD: Ryzen 3/5/7/9 + модель
        Ryzen\s+(?:3|5|7|9)\s+\d+[A-Z]*
        |
        # AMD: Threadripper + модель
        Threadripper\s+\d+[A-Z]*
        |
        # AMD: Athlon + модель
        Athlon(?:\s+(?:Gold|Silver))?\s+\d+[A-Z]*
    )
    """,
    re.IGNORECASE | re.VERBOSE | re.UNICODE,
)


def normalize_cpu_model(cpu_str: str) -> str:
    """Нормалізує модель процесора до стандартного формату.

    Видаляє зайві пробіли, приводить виробника до стандартного регістру.

    Аргументи:
        cpu_str: Оригінальний рядок з моделлю процесора.

    Повертає:
        Нормалізований рядок (наприклад, "Intel Core i7-12700F").

    Приклад:
        normalize_cpu_model("intel core i7-12700f")  →  "Intel Core i7-12700F"
        normalize_cpu_model("AMD ryzen 5 7520U")     →  "AMD Ryzen 5 7520U"
    """
    # Видаляємо зайві пробіли
    cpu_str = " ".join(cpu_str.split())
    
    # Приводимо виробника до стандартного регістру
    if cpu_str.lower().startswith("intel"):
        cpu_str = "Intel" + cpu_str[5:]
    elif cpu_str.lower().startswith("amd"):
        cpu_str = "AMD" + cpu_str[3:]
    
    # Нормалізуємо "Core i7", "Ryzen 5", тощо
    cpu_str = re.sub(r'\bcore\b', 'Core', cpu_str, flags=re.IGNORECASE)
    cpu_str = re.sub(r'\bpentium\b', 'Pentium', cpu_str, flags=re.IGNORECASE)
    cpu_str = re.sub(r'\bceleron\b', 'Celeron', cpu_str, flags=re.IGNORECASE)
    cpu_str = re.sub(r'\bryzen\b', 'Ryzen', cpu_str, flags=re.IGNORECASE)
    cpu_str = re.sub(r'\bathlon\b', 'Athlon', cpu_str, flags=re.IGNORECASE)
    cpu_str = re.sub(r'\bthreadripper\b', 'Threadripper', cpu_str, flags=re.IGNORECASE)
    cpu_str = re.sub(r'\bultra\b', 'Ultra', cpu_str, flags=re.IGNORECASE)
    cpu_str = re.sub(r'\bgold\b', 'Gold', cpu_str, flags=re.IGNORECASE)
    cpu_str = re.sub(r'\bsilver\b', 'Silver', cpu_str, flags=re.IGNORECASE)
    cpu_str = re.sub(r'\bprocessor\b', 'processor', cpu_str, flags=re.IGNORECASE)
    
    return cpu_str


def extract_cpu_model_matches(text: str) -> list[tuple[str, str]]:
    """Знаходить усі моделі процесорів у тексті.

    Аргументи:
        text: Рядок для пошуку.

    Повертає:
        Список пар (оригінал, нормалізована_модель).
        Порожній список, якщо моделей не знайдено.

    Приклад:
        extract_cpu_model_matches("Ноутбук з Intel Core i7-12700F процесором")
        →  [('Intel Core i7-12700F', 'Intel Core i7-12700F')]
    """
    if not isinstance(text, str) or not text.strip():
        return []

    results: list[tuple[str, str]] = []
    seen: set[str] = set()

    for match in _CPU_MODEL_RE.finditer(text):
        original = match.group(0)
        normalized = normalize_cpu_model(original)
        
        if normalized not in seen:
            results.append((original, normalized))
            seen.add(normalized)

    return results


# ---------------------------------------------------------------------------
# Модель відеокарти (GPU)
# ---------------------------------------------------------------------------

# Регулярний вираз для пошуку моделей відеокарт.
# Підтримує:
#   nVidia: GeForce RTX/GTX, Quadro
#   AMD: Radeon RX/Pro
#   Intel: Arc, HD Graphics, UHD Graphics, Iris
# Приклади:
#   "GeForce RTX 5060 Ti", "Radeon RX 9070 XT", "Intel Arc Graphics"
_GPU_MODEL_RE = re.compile(
    r"""
    (?:
        # nVidia GeForce
        (?:nVidia\s+)?GeForce\s+(?:RTX|GTX)\s+\d+(?:\s+Ti)?
        |
        # nVidia Quadro
        (?:nVidia\s+)?Quadro\s+[A-Z]*\d+
        |
        # AMD Radeon
        (?:AMD\s+)?Radeon\s+(?:RX|Pro)\s+\d+(?:\s+XT)?
        |
        # Intel Arc
        Intel\s+Arc\s+(?:Graphics|[A-Z]\d+)
        |
        # Intel HD/UHD Graphics
        Intel\s+(?:HD|UHD)\s+Graphics(?:\s+\d+)?
        |
        # Intel Iris
        Intel\s+Iris(?:\s+(?:Xe|Plus|Pro))?\s*(?:Graphics)?(?:\s+\d+)?
    )
    """,
    re.IGNORECASE | re.VERBOSE | re.UNICODE,
)


def normalize_gpu_model(gpu_str: str) -> str:
    """Нормалізує модель відеокарти до стандартного формату.

    Аргументи:
        gpu_str: Оригінальний рядок з моделлю відеокарти.

    Повертає:
        Нормалізований рядок.

    Приклад:
        normalize_gpu_model("geforce rtx 5060 ti")  →  "GeForce RTX 5060 Ti"
        normalize_gpu_model("radeon rx 9070 xt")    →  "Radeon RX 9070 XT"
        normalize_gpu_model("nvidia geforce rtx 5060 ti")  →  "GeForce RTX 5060 Ti"
    """
    # Видаляємо зайві пробіли
    gpu_str = " ".join(gpu_str.split())
    
    # Видаляємо префікси виробника (nVidia, AMD) якщо вони є перед назвою серії
    gpu_str = re.sub(r'\bnvidia\s+', '', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bamd\s+(?=radeon)', '', gpu_str, flags=re.IGNORECASE)
    
    # Нормалізуємо назви брендів та серій
    gpu_str = re.sub(r'\bgeforce\b', 'GeForce', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\brtx\b', 'RTX', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bgtx\b', 'GTX', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bquadro\b', 'Quadro', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bradeon\b', 'Radeon', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\brx\b', 'RX', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bpro\b', 'Pro', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bintel\b', 'Intel', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\barc\b', 'Arc', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bhd\b', 'HD', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\buhd\b', 'UHD', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\biris\b', 'Iris', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bgraphics\b', 'Graphics', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bti\b', 'Ti', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bxt\b', 'XT', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bxe\b', 'Xe', gpu_str, flags=re.IGNORECASE)
    gpu_str = re.sub(r'\bplus\b', 'Plus', gpu_str, flags=re.IGNORECASE)
    
    return gpu_str


def extract_gpu_model_matches(text: str) -> list[tuple[str, str]]:
    """Знаходить усі моделі відеокарт у тексті.

    Аргументи:
        text: Рядок для пошуку.

    Повертає:
        Список пар (оригінал, нормалізована_модель).
        Порожній список, якщо моделей не знайдено.

    Приклад:
        extract_gpu_model_matches("nVidia GeForce RTX 5060 Ti, 16 ГБ")
        →  [('GeForce RTX 5060 Ti', 'GeForce RTX 5060 Ti')]
    """
    if not isinstance(text, str) or not text.strip():
        return []

    results: list[tuple[str, str]] = []
    seen: set[str] = set()

    for match in _GPU_MODEL_RE.finditer(text):
        original = match.group(0)
        normalized = normalize_gpu_model(original)
        
        if normalized not in seen:
            results.append((original, normalized))
            seen.add(normalized)

    return results


# ---------------------------------------------------------------------------
# Операційна система (OS)
# ---------------------------------------------------------------------------

# Регулярний вираз для пошуку операційних систем.
# Підтримує: Windows, Linux, macOS, DOS
_OS_RE = re.compile(
    r"""
    (?:
        # Windows з версією
        Windows\s+(?:
            11|10|8\.1|8|7|Vista|XP|2000
        )(?:\s+(?:Home|Pro|Enterprise|Education|S))?
        |
        # macOS
        macOS(?:\s+(?:Ventura|Monterey|Big\s+Sur|Catalina|Mojave|High\s+Sierra|Sierra))?
        |
        # Linux дистрибутиви
        (?:Ubuntu|Debian|Fedora|CentOS|Red\s+Hat|Arch|Linux)(?:\s+\d+(?:\.\d+)?)?
        |
        # FreeDOS / DOS
        (?:Free)?DOS
        |
        # Без ОС
        без\s+О[СД]|без\s+операц[іи]йно[їі]\s+систем[иы]
    )
    """,
    re.IGNORECASE | re.VERBOSE | re.UNICODE,
)


def normalize_os(os_str: str) -> str:
    """Нормалізує назву операційної системи до стандартного формату.

    Аргументи:
        os_str: Оригінальний рядок з назвою ОС.

    Повертає:
        Нормалізований рядок.

    Приклад:
        normalize_os("windows 11 home")  →  "Windows 11 Home"
        normalize_os("MACOS")            →  "macOS"
    """
    # Видаляємо зайві пробіли
    os_str = " ".join(os_str.split())
    
    # Нормалізуємо Windows
    os_str = re.sub(r'\bwindows\b', 'Windows', os_str, flags=re.IGNORECASE)
    os_str = re.sub(r'\bhome\b', 'Home', os_str, flags=re.IGNORECASE)
    os_str = re.sub(r'\bpro\b', 'Pro', os_str, flags=re.IGNORECASE)
    os_str = re.sub(r'\benterprise\b', 'Enterprise', os_str, flags=re.IGNORECASE)
    os_str = re.sub(r'\beducation\b', 'Education', os_str, flags=re.IGNORECASE)
    
    # Нормалізуємо macOS
    if os_str.lower().startswith('macos'):
        os_str = 'macOS' + os_str[5:]
    
    # Нормалізуємо Linux
    os_str = re.sub(r'\bubuntu\b', 'Ubuntu', os_str, flags=re.IGNORECASE)
    os_str = re.sub(r'\bdebian\b', 'Debian', os_str, flags=re.IGNORECASE)
    os_str = re.sub(r'\bfedora\b', 'Fedora', os_str, flags=re.IGNORECASE)
    os_str = re.sub(r'\blinux\b', 'Linux', os_str, flags=re.IGNORECASE)
    
    # Нормалізуємо DOS
    os_str = re.sub(r'\b(?:free)?dos\b', 'DOS', os_str, flags=re.IGNORECASE)
    if 'free' in os_str.lower() and 'dos' in os_str.lower():
        os_str = 'FreeDOS'
    
    return os_str


def extract_os_matches(text: str) -> list[tuple[str, str]]:
    """Знаходить усі операційні системи у тексті.

    Аргументи:
        text: Рядок для пошуку.

    Повертає:
        Список пар (оригінал, нормалізована_ОС).
        Порожній список, якщо ОС не знайдено.

    Приклад:
        extract_os_matches("Ноутбук з Windows 11 Home")
        →  [('Windows 11 Home', 'Windows 11 Home')]
    """
    if not isinstance(text, str) or not text.strip():
        return []

    results: list[tuple[str, str]] = []
    seen: set[str] = set()

    for match in _OS_RE.finditer(text):
        original = match.group(0)
        normalized = normalize_os(original)
        
        if normalized not in seen:
            results.append((original, normalized))
            seen.add(normalized)

    return results


# ===========================================================================
# Модуль 2c: Допоміжні функції витягування канонічного значення та диспетчер
# ===========================================================================


# ---------------------------------------------------------------------------
# Числове порівняння з урахуванням одиниць та виявлення опечаток
# ---------------------------------------------------------------------------


def _norm_to_base_value(normalized: str, checker_type: str) -> float | None:
    """Перетворює нормалізоване значення до базової числової одиниці.

    Використовується для порівняння значень у різних одиницях виміру.

    Аргументи:
        normalized:   Нормалізований рядок (напр. «512ГБ», «1.5кг», «5000мАг»).
        checker_type: Тип перевірки (з VALIDATION_RULES).

    Повертає:
        float або None.
    """
    if checker_type == "memory":
        return memory_to_mb(normalized)
    if checker_type == "battery":
        return battery_to_mah(normalized)
    if checker_type == "weight":
        return weight_to_grams(normalized)
    if checker_type == "screen_diagonal":
        m = re.match(r"^(\d+(?:\.\d+)?)", normalized)
        return float(m.group(1)) if m else None
    return None


def _numeric_values_equal(norm1: str, norm2: str, checker_type: str) -> bool:
    """Перевіряє числову еквівалентність двох значень з можливою конвертацією одиниць.

    Наприклад: «1ТБ» == «1024ГБ», «5Аг» == «5000мАг», «1.5кг» == «1500г».

    Аргументи:
        norm1:        Перше нормалізоване значення.
        norm2:        Друге нормалізоване значення.
        checker_type: Тип перевірки — визначає алгоритм конвертації.

    Повертає:
        True, якщо значення чисельно еквівалентні.
    """
    if norm1 == norm2:
        return True
    val1 = _norm_to_base_value(norm1, checker_type)
    val2 = _norm_to_base_value(norm2, checker_type)
    if val1 is not None and val2 is not None and val1 > 0:
        return abs(val1 - val2) / val1 < 1e-6
    return False


def _get_typo_hint(
    canonical_norm: str, found_norms: list[str], checker_type: str
) -> str | None:
    """Повертає підказку, якщо значення відрізняються на типовий «зайвий нуль» (×10 або ×100).

    Аргументи:
        canonical_norm: Нормалізоване еталонне значення.
        found_norms:    Список нормалізованих значень, знайдених у тексті.
        checker_type:   Тип перевірки.

    Повертає:
        Рядок-підказку або None.

    Приклад:
        _get_typo_hint("512мАг", ["5120мАг"], "battery")
        →  "можливо зайвий нуль (різниця ×10)"
    """
    canonical_val = _norm_to_base_value(canonical_norm, checker_type)
    if canonical_val is None or canonical_val == 0:
        return None
    for norm in found_norms:
        found_val = _norm_to_base_value(norm, checker_type)
        if found_val is None or found_val == 0:
            continue
        ratio = max(found_val, canonical_val) / min(found_val, canonical_val)
        for factor in (10, 100):
            if abs(ratio - factor) / factor < 0.15:
                return f"можливо зайвий нуль (різниця ×{factor})"
    return None


def _normalize_canonical(raw: str, rule: dict) -> str | None:
    """Витягує нормалізоване канонічне значення з колонки характеристики.

    Для кожного типу перевірки (checker_type) застосовує відповідний алгоритм.
    Якщо колонка характеристики містить значення без одиниці виміру (наприклад,
    "15.6" замість "15.6""), намагається витягти числову частину напряму.

    Аргументи:
        raw:  Рядок зі значенням колонки характеристики (після str().strip()).
        rule: Словник правила з VALIDATION_RULES.

    Повертає:
        Нормалізований рядок або None, якщо значення не вдалось розпізнати.
    """
    checker_type = rule.get("checker_type", "memory")

    if checker_type == "memory":
        matches = extract_memory_matches(raw)
        return matches[0][1] if matches else None

    if checker_type == "screen_diagonal":
        matches = extract_screen_diagonal_matches(raw)
        if matches:
            return matches[0][1]
        # Колонка характеристики може містити просте число: "15.6"
        bare = re.match(r"^\s*(\d+(?:[.,]\d+)?)\s*$", raw)
        if bare:
            number = bare.group(1).replace(",", ".")
            if "." in number:
                number = number.rstrip("0").rstrip(".")
            return f'{number}"'
        return None

    if checker_type == "weight":
        matches = extract_weight_matches(raw)
        if matches:
            return matches[0][1]
        # Колонка характеристики може містити просте число: "1.5" (вважаємо кг)
        bare = re.match(r"^\s*(\d+(?:[.,]\d+)?)\s*$", raw)
        if bare:
            number = bare.group(1).replace(",", ".")
            if "." in number:
                number = number.rstrip("0").rstrip(".")
            return f"{number}кг"
        return None

    if checker_type == "battery":
        matches = extract_battery_matches(raw)
        if matches:
            return matches[0][1]
        # Колонка характеристики може містити просте число (вважаємо мАг)
        bare = re.match(r"^\s*(\d+(?:[.,]\d+)?)\s*$", raw)
        if bare:
            number = bare.group(1).replace(",", ".")
            if "." in number:
                number = number.rstrip("0").rstrip(".")
            return f"{number}мАг"
        return None

    if checker_type == "resolution":
        raw_lower = raw.strip().lower()
        # Перевіряємо псевдоніми
        if raw_lower in _RESOLUTION_ALIASES:
            return _RESOLUTION_ALIASES[raw_lower]
        # Перевіряємо числовий формат
        m = _RESOLUTION_NUM_RE.search(raw)
        if m:
            return f"{m.group(1)}x{m.group(2)}"
        return None

    if checker_type == "value_list":
        valid_values = rule.get("valid_values", [])
        raw_lower = raw.strip().lower()
        # Точний збіг (нечутливий до регістру)
        for val in valid_values:
            if val.lower() == raw_lower:
                return val.upper()
        # Часткове входження (для значень типу "Дискретна відеокарта")
        for val in sorted(valid_values, key=len, reverse=True):
            if val.lower() in raw_lower:
                return val.upper()
        return None

    if checker_type == "cpu_model":
        matches = extract_cpu_model_matches(raw)
        return matches[0][1] if matches else None

    if checker_type == "gpu_model":
        matches = extract_gpu_model_matches(raw)
        return matches[0][1] if matches else None

    if checker_type == "os":
        matches = extract_os_matches(raw)
        return matches[0][1] if matches else None

    return None


def _extract_for_rule(text: str, rule: dict) -> list[tuple[str, str]]:
    """Знаходить у тексті значення відповідного типу характеристики.

    Диспетчер: обирає функцію-екстрактор залежно від поля "checker_type" правила.

    Аргументи:
        text: Рядок для пошуку (вже очищений від HTML).
        rule: Словник правила з VALIDATION_RULES.

    Повертає:
        Список пар (оригінал, нормалізоване). Порожній список, якщо нічого не знайдено.
    """
    checker_type = rule.get("checker_type", "memory")

    if checker_type == "memory":
        # Використовуємо context-aware extraction якщо є ключові слова
        context_keywords = rule.get("context_keywords")
        return extract_memory_matches_with_context(text, context_keywords)
    if checker_type == "screen_diagonal":
        return extract_screen_diagonal_matches(text)
    if checker_type == "weight":
        return extract_weight_matches(text)
    if checker_type == "battery":
        return extract_battery_matches(text)
    if checker_type == "resolution":
        return extract_resolution_matches(text)
    if checker_type == "value_list":
        return extract_value_list_matches(text, rule.get("valid_values", []))
    if checker_type == "cpu_model":
        return extract_cpu_model_matches(text)
    if checker_type == "gpu_model":
        return extract_gpu_model_matches(text)
    if checker_type == "os":
        return extract_os_matches(text)

    return []


# ===========================================================================
# Модуль 4: Правила валідації (імпортуються з validation_rules.py)
# ===========================================================================
# VALIDATION_RULES імпортовано на початку файлу: from validation_rules import VALIDATION_RULES


# ===========================================================================
# Модуль 5: Логіка порівняння та виявлення конфліктів
# ===========================================================================


def check_conflicts(
    row: pd.Series,
    df_columns: list[str],
    rule: dict,
) -> tuple[str, list[str]]:
    """Перевіряє наявність конфліктів для одного правила в одному рядку товару.

    Алгоритм:
        1. Знаходить колонку з еталонним значенням характеристики.
        2. Витягує та нормалізує еталонне значення залежно від типу checker_type.
        3. Для кожного текстового поля (після очищення від HTML) знаходить
           значення того ж типу за допомогою відповідного екстрактора.
        4. Якщо у текстовому полі є значення, відмінні від еталонного —
           це конфлікт.
        5. Якщо текстове поле не містить жодного значення — пропускає
           (відсутність згадки не є конфліктом).

    Аргументи:
        row:        Рядок товару (pd.Series).
        df_columns: Список усіх назв колонок DataFrame.
        rule:       Словник правила з VALIDATION_RULES.

    Повертає:
        Кортеж (рядок_помилки, список_конфліктних_колонок).
        Якщо конфліктів немає — ("", []).

    Приклад:
        При Название="Ноутбук 128ГБ SSD", Характеристика SSD="512 ГБ":
        → ("Конфлікт SSD: Назва (128ГБ) != Характеристика ... (512 ГБ)",
           ["Название", "Объём SSD;115411"])
    """
    label = rule["label"]

    # Крок 1: Знаходимо колонку з еталонним значенням характеристики
    char_col: str | None = None
    for hint in rule["char_hints"]:
        char_col = find_column(df_columns, hint)
        if char_col:
            break
    if char_col is None:
        # Колонка характеристики відсутня у файлі — пропускаємо правило
        return "", []

    # Крок 2: Отримуємо та нормалізуємо еталонне значення
    raw_char = row.get(char_col, "")
    if pd.isna(raw_char) or not str(raw_char).strip():
        return "", []  # Порожня характеристика — немає з чим порівнювати

    canonical_raw = str(raw_char).strip()
    canonical_norm = _normalize_canonical(canonical_raw, rule)
    if canonical_norm is None:
        # Характеристика заповнена, але не містить розпізнаного значення
        return "", []

    checker_type = rule.get("checker_type", "memory")

    # Крок 3: Перевіряємо кожне текстове поле
    conflict_parts: list[str] = []   # Частини тексту повідомлення про конфлікт
    conflict_cols: list[str] = []    # Назви колонок, де виявлено конфлікт
    seen_text_cols: set[str] = set() # Щоб не перевіряти одну й ту ж колонку двічі

    for hint in rule["text_hints"]:
        text_col = find_column(df_columns, hint)
        if text_col is None or text_col in seen_text_cols:
            continue  # Текстова колонка відсутня або вже оброблена
        seen_text_cols.add(text_col)

        raw_text = row.get(text_col, "")
        if pd.isna(raw_text) or not str(raw_text).strip():
            continue  # Порожнє поле — не конфлікт

        # Очищаємо HTML та шукаємо значення відповідного типу
        clean_text = clean_html(str(raw_text))
        matches = _extract_for_rule(clean_text, rule)

        if not matches:
            continue  # Поле не містить згадки про цю характеристику — не конфлікт

        # Крок 4: Конфлікт є, якщо канонічного значення серед знайдених немає
        found_norms = [norm for _, norm in matches]
        if canonical_norm not in found_norms:
            # Перевіряємо числову еквівалентність з конвертацією одиниць
            # (наприклад, 1ТБ == 1024ГБ, 5Аг == 5000мАг, 1.5кг == 1500г)
            if any(
                _numeric_values_equal(canonical_norm, n, checker_type)
                for n in found_norms
            ):
                continue  # Значення чисельно еквівалентні — конфлікту немає

            found_raws = [raw for raw, _ in matches]
            typo_hint = _get_typo_hint(canonical_norm, found_norms, checker_type)
            msg = f"{text_col} ({', '.join(found_raws)})"
            if typo_hint:
                msg += f" [{typo_hint}]"
            conflict_parts.append(msg)
            conflict_cols.append(text_col)

    # Крок 5: Формуємо повідомлення про конфлікт
    if conflict_parts:
        char_part = f"Характеристика {char_col} ({canonical_raw})"
        error_msg = (
            f"Конфлікт {label}: "
            f"{' != '.join(conflict_parts)} != {char_part}"
        )
        conflict_cols.append(char_col)
        return error_msg, conflict_cols

    return "", []


def check_language_mismatch(
    row: pd.Series,
    df_columns: list[str],
) -> tuple[str, list[str]]:
    """Перевіряє відповідність мов у парах російський/український текст.
    
    Перевіряє наступні пари колонок:
    - "Название" (російська) та "Назва (ua)" (українська)
    - Характеристики без "(ua)" (російські) та з "(ua)" (українські)
    
    Аргументи:
        row:        Рядок товару (pd.Series).
        df_columns: Список усіх назв колонок DataFrame.
    
    Повертає:
        Кортеж (рядок_помилки, список_конфліктних_колонок).
        Якщо мови відповідають — ("", []).
    
    Приклад:
        При Название="Ноутбук із SSD", Назва (ua)="Ноутбук с SSD":
        → ("Невідповідність мови: Название має бути російською, Назва (ua) має бути українською",
           ["Название", "Назва (ua)"])
    """
    errors = []
    conflict_cols = []
    
    # Перевірка 1: Назва товару (Название / Назва (ua))
    ru_name_col = find_column(df_columns, "Название")
    uk_name_col = find_column(df_columns, "Назва (ua)")
    
    if ru_name_col and uk_name_col:
        ru_text = str(row.get(ru_name_col, "")).strip()
        uk_text = str(row.get(uk_name_col, "")).strip()
        
        # Очищаємо від HTML
        ru_text = clean_html(ru_text) if ru_text else ""
        uk_text = clean_html(uk_text) if uk_text else ""
        
        if ru_text or uk_text:
            is_valid, error_msg = check_language_consistency(ru_text, uk_text, allow_empty=True)
            if not is_valid:
                errors.append(f"Назва: {error_msg}")
                if ru_text:
                    conflict_cols.append(ru_name_col)
                if uk_text:
                    conflict_cols.append(uk_name_col)
    
    # Перевірка 2: Характеристики (знаходимо пари російська/українська)
    # Групуємо колонки за базовою назвою (до ";")
    char_pairs: dict[str, dict[str, str]] = {}  # базова назва → {"ru": col, "uk": col}
    
    for col in df_columns:
        base_name = col.split(";")[0].strip()
        
        # Пропускаємо колонки назв та описів (вони не є характеристиками)
        if base_name in ["Название", "Назва (ua)", "Краткое описание", "Артикул"]:
            continue
        if "Описание" in base_name or "Опис" in base_name:
            continue
        
        # Визначаємо, чи це українська версія
        if base_name.endswith(" (ua)"):
            # Українська версія
            ru_base = base_name[:-5].strip()  # Прибираємо " (ua)"
            if ru_base not in char_pairs:
                char_pairs[ru_base] = {}
            char_pairs[ru_base]["uk"] = col
        else:
            # Російська версія (або нейтральна)
            if base_name not in char_pairs:
                char_pairs[base_name] = {}
            char_pairs[base_name]["ru"] = col
    
    # Перевіряємо кожну пару
    for base_name, pair in char_pairs.items():
        if "ru" in pair and "uk" in pair:
            ru_col = pair["ru"]
            uk_col = pair["uk"]
            
            ru_text = str(row.get(ru_col, "")).strip()
            uk_text = str(row.get(uk_col, "")).strip()
            
            # Очищаємо від HTML
            ru_text = clean_html(ru_text) if ru_text else ""
            uk_text = clean_html(uk_text) if uk_text else ""
            
            if ru_text or uk_text:
                is_valid, error_msg = check_language_consistency(ru_text, uk_text, allow_empty=True)
                if not is_valid:
                    errors.append(f"Характеристика '{base_name}': {error_msg}")
                    if ru_text:
                        conflict_cols.append(ru_col)
                    if uk_text:
                        conflict_cols.append(uk_col)
    
    # Формуємо фінальне повідомлення
    if errors:
        error_msg = "Невідповідність мови: " + " | ".join(errors)
        return error_msg, conflict_cols
    
    return "", []


# ===========================================================================
# Модуль 5c: Семантична перевірка
# ===========================================================================

# Пари семантично протилежних термінів.
# Якщо терміни group_a знайдено в одній колонці, а group_b — в іншій,
# це вважається суперечністю.
_SEMANTIC_CONTRADICTIONS: list[dict] = [
    {
        "label": "Бездротовий/Дротовий",
        "group_a": [
            "бездротовий", "бездротова", "бездротове", "бездротові",
            "беспроводной", "беспроводная", "беспроводное",
            "wireless", "wi-fi", "wifi", "bluetooth",
        ],
        "group_b": [
            "дротовий", "дротова", "дротове", "дротові",
            "проводной", "проводная", "проводное",
            "з дротом", "с проводом", "wired",
        ],
    },
    {
        "label": "Підсвітка клавіатури",
        "group_a": [
            "з підсвіткою клавіатури", "підсвітка клавіатури",
            "backlit keyboard", "keyboard backlit",
            "с подсветкой клавиатуры",
        ],
        "group_b": [
            "без підсвітки клавіатури", "без підсвітки",
            "без подсветки клавиатуры", "без подсветки",
        ],
    },
    {
        "label": "Сенсорний екран",
        "group_a": [
            "сенсорний екран", "сенсорний дисплей",
            "touchscreen", "touch screen",
            "сенсорный экран", "сенсорный дисплей",
        ],
        "group_b": [
            "несенсорний", "не сенсорний",
            "несенсорный", "без сенсорного",
            "non-touch",
        ],
    },
    {
        "label": "Трансформер/Класичний",
        "group_a": [
            "трансформер", "2-в-1", "2 в 1", "two-in-one", "2-in-1",
            "конвертований", "планшет-трансформер",
        ],
        "group_b": [
            "не трансформер", "звичайний ноутбук", "класичний ноутбук",
        ],
    },
]

# Попередньо скомпільовані патерни для кожного терміна з _SEMANTIC_CONTRADICTIONS.
# Ключ: термін (нижній регістр), значення: скомпільований re.Pattern.
# Зберігаємо тут, щоб не перекомпілювати на кожному виклику check_semantic_conflicts.
_SEMANTIC_TERM_PATTERNS: dict[str, re.Pattern] = {}
for _contradiction in _SEMANTIC_CONTRADICTIONS:
    for _term in itertools.chain(_contradiction["group_a"], _contradiction["group_b"]):
        _key = _term.lower()
        if _key not in _SEMANTIC_TERM_PATTERNS:
            _SEMANTIC_TERM_PATTERNS[_key] = re.compile(
                r"\b" + re.escape(_key) + r"\b", re.UNICODE
            )

# Регулярний вираз для кількості процесорних ядер.
# Розпізнає: «12-ядерний», «4 ядерна», «8 cores», «6-core» тощо.
_CORE_COUNT_RE = re.compile(
    r"(\d+)\s*[-–—]?\s*"
    r"(?:ядерн(?:ий|а|е|і|их|ій|ой|ая|ое)|ядер(?:ний)?|ядро|"
    r"core(?:s)?)",
    re.IGNORECASE | re.UNICODE,
)

# Колонки, у яких шукаємо семантичні суперечності
_SEMANTIC_TEXT_HINTS = ["Название", "Назва", "Краткое описание", "Описание"]


def check_semantic_conflicts(
    row: pd.Series,
    df_columns: list[str],
) -> tuple[str, list[str]]:
    """Перевіряє семантичні суперечності між текстовими полями товару.

    Виявляє:
      1. Протилежні за змістом терміни в різних колонках:
         наприклад, «бездротовий» у назві, але «з дротом» в описі.
      2. Суперечності у кількості ядер процесора:
         наприклад, «12-ядерний» у назві, але «4-ядерний» в описі.

    Аргументи:
        row:        Рядок товару (pd.Series).
        df_columns: Список усіх назв колонок DataFrame.

    Повертає:
        Кортеж (рядок_помилки, список_конфліктних_колонок).
        Якщо суперечностей немає — ("", []).
    """
    # Збираємо вміст усіх текстових колонок
    text_col_contents: dict[str, str] = {}
    for hint in _SEMANTIC_TEXT_HINTS:
        col = find_column(df_columns, hint)
        if col and col not in text_col_contents:
            raw = row.get(col, "")
            if not pd.isna(raw) and str(raw).strip():
                text_col_contents[col] = clean_html(str(raw)).lower()

    if not text_col_contents:
        return "", []

    def _term_matches(term: str, text: str) -> bool:
        """Перевіряє наявність терміна у тексті як цілого слова (з межами слова)."""
        pattern = _SEMANTIC_TERM_PATTERNS.get(term)
        if pattern is None:
            # Запасний варіант: компіляція на льоту (для непередбачених термінів)
            pattern = re.compile(r"\b" + re.escape(term) + r"\b", re.UNICODE)
        return bool(pattern.search(text))

    errors: list[str] = []
    conflict_cols: list[str] = []

    # Перевірка 1: Пари семантично протилежних термінів
    for contradiction in _SEMANTIC_CONTRADICTIONS:
        label = contradiction["label"]
        group_a = [t.lower() for t in contradiction["group_a"]]
        group_b = [t.lower() for t in contradiction["group_b"]]

        cols_with_a: list[str] = []
        cols_with_b: list[str] = []

        for col, text in text_col_contents.items():
            for term in group_a:
                if _term_matches(term, text):
                    cols_with_a.append(col)
                    break
            for term in group_b:
                if _term_matches(term, text):
                    cols_with_b.append(col)
                    break

        if cols_with_a and cols_with_b:
            errors.append(
                f"Семантичний конфлікт '{label}': "
                f"протиріччя між {', '.join(cols_with_a)} та {', '.join(cols_with_b)}"
            )
            for col in set(cols_with_a + cols_with_b):
                if col not in conflict_cols:
                    conflict_cols.append(col)

    # Перевірка 2: Кількість ядер процесора
    core_counts: dict[str, set[str]] = {}
    for col, text in text_col_contents.items():
        found = set(_CORE_COUNT_RE.findall(text))
        if found:
            core_counts[col] = found

    if len(core_counts) >= 2:
        all_counts: set[str] = set()
        for counts in core_counts.values():
            all_counts.update(counts)
        if len(all_counts) > 1:
            details = "; ".join(
                f"{col}: {', '.join(sorted(counts))}-ядерний"
                for col, counts in core_counts.items()
            )
            errors.append(f"Конфлікт кількості ядер: {details}")
            for col in core_counts:
                if col not in conflict_cols:
                    conflict_cols.append(col)

    if errors:
        return "Семантична перевірка: " + " | ".join(errors), conflict_cols

    return "", []


# ===========================================================================
# Модуль 6: Читання вхідного файлу
# ===========================================================================


def read_input_file(filepath: str | Path) -> pd.DataFrame:
    """Зчитує вхідний файл (CSV або Excel) у pandas DataFrame.

    Підтримує формати: .csv, .xlsx, .xls.
    Усі значення читаються як рядки (dtype=str), щоб уникнути небажаних
    автоматичних перетворень типів.
    Для CSV автоматично пробує роздільники: кома, крапка з комою, табуляція;
    обирає той, що дає найбільшу кількість колонок.

    Аргументи:
        filepath: Шлях до вхідного файлу.

    Повертає:
        pandas DataFrame із даними файлу.

    Виключення:
        FileNotFoundError: Якщо файл не знайдено за вказаним шляхом.
        ValueError:        Якщо формат файлу не підтримується.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Файл не знайдено: {filepath}")

    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        return pd.read_excel(path, dtype=str)

    if suffix == ".csv":
        best_df: pd.DataFrame | None = None
        for sep in (",", ";", "\t"):
            try:
                df = pd.read_csv(path, sep=sep, dtype=str, encoding="utf-8-sig")
                if best_df is None or len(df.columns) > len(best_df.columns):
                    best_df = df
            except Exception:
                continue
        if best_df is not None:
            return best_df
        # Останній варіант: автовизначення роздільника через pandas
        return pd.read_csv(
            path, sep=None, engine="python", dtype=str, encoding="utf-8-sig"
        )

    raise ValueError(
        f"Непідтримуваний формат файлу: '{suffix}'. Підтримуються: .csv, .xlsx, .xls"
    )


# ===========================================================================
# Модуль 7: Збереження результатів у Excel із підсвіткою
# ===========================================================================


def _save_with_highlights(
    df: pd.DataFrame,
    output_file: str,
    conflict_cells: dict[tuple[int, str], bool],
) -> None:
    """Зберігає DataFrame у .xlsx та підсвічує конфліктні клітинки.

    Алгоритм:
        1. Зберігає DataFrame через pandas (engine=openpyxl) без форматування.
        2. Відкриває збережений файл через openpyxl для ручного форматування.
        3. Задає червоний фон для клітинок із конфліктними значеннями.
        4. Задає жовтий фон для клітинок колонки "Помилки".
        5. Зберігає відформатований файл.

    Примітка:
        Звичайний pandas не підтримує форматування клітинок. Для підсвітки
        кольором необхідний openpyxl. Спочатку зберігаємо через pandas
        (швидко), потім відкриваємо через openpyxl лише для форматування.

    Аргументи:
        df:             DataFrame для збереження.
        output_file:    Шлях до вихідного .xlsx файлу.
        conflict_cells: Словник {(df_row_index, col_name): True} —
                        координати конфліктних клітинок.
    """
    # Крок 1: Зберігаємо без форматування
    df.to_excel(output_file, index=False, engine="openpyxl")

    # Крок 2: Відкриваємо для форматування
    wb = load_workbook(output_file)
    ws = wb.active

    # Будуємо маппінг: назва_колонки → номер_колонки Excel (1-based)
    header_to_col: dict[str, int] = {
        cell.value: cell.column for cell in ws[1] if cell.value is not None
    }

    # Будуємо маппінг: індекс рядка DataFrame → номер рядка Excel
    # (рядок 1 — заголовки, дані починаються з рядка 2)
    df_idx_to_excel_row: dict[int, int] = {
        df_idx: excel_row
        for excel_row, df_idx in enumerate(df.index, start=2)
    }

    # Крок 3: Підсвічуємо кожну конфліктну клітинку
    for (df_idx, col_name) in conflict_cells:
        excel_row = df_idx_to_excel_row.get(df_idx)
        excel_col = header_to_col.get(col_name)
        if excel_row is None or excel_col is None:
            continue

        cell = ws.cell(row=excel_row, column=excel_col)
        if col_name == "Помилки":
            cell.fill = ERROR_FILL    # Жовтий — для колонки з описом помилок
        else:
            cell.fill = CONFLICT_FILL  # Червоний — для полів із конфліктами

    # Крок 4: Зберігаємо відформатований файл
    wb.save(output_file)


# ===========================================================================
# Модуль 8: Головна функція валідації
# ===========================================================================


def validate_feed(input_file: str, output_file: str) -> pd.DataFrame:
    """Головна функція: зчитує товарний фід, валідує та зберігає результат.

    Алгоритм:
        1. Зчитує вхідний файл у DataFrame.
        2. Визначає, які правила застосовуються до цього файлу
           (яких колонок характеристик є у файлі).
        3. Для кожного рядка та кожного активного правила перевіряє
           наявність конфліктів між текстовими полями та характеристикою.
        4. Записує знайдені помилки у нову колонку "Помилки".
        5. Зберігає result.xlsx із підсвіченими конфліктними клітинками
           (червоний — конфліктні поля, жовтий — опис помилки).
        6. Виводить повний звіт: які характеристики перевірено,
           скільки конфліктів знайдено по кожній.

    Аргументи:
        input_file:  Шлях до вхідного файлу (.csv, .xlsx, .xls).
        output_file: Шлях до вихідного файлу (.xlsx).

    Повертає:
        DataFrame із доданою колонкою "Помилки".

    Виключення:
        FileNotFoundError: Якщо input_file не знайдено.
        ValueError:        Якщо формат файлу не підтримується.

    Приклад:
        df = validate_feed("products.csv", "result.xlsx")
    """
    print(f"Читаємо файл: {input_file}")
    df = read_input_file(input_file)
    print(f"Зчитано {len(df)} рядків, {len(df.columns)} колонок.")
    df_columns = list(df.columns)

    # Додаємо колонку для помилок, якщо її ще немає
    if "Помилки" not in df.columns:
        df["Помилки"] = ""

    # Словник координат конфліктних клітинок:
    # ключ = (df_row_index, column_name), значення = True (для унікальності)
    conflict_cells: dict[tuple[int, str], bool] = {}

    # Лічильники для повного звіту: label → {"applicable": bool, "conflicts": int}
    rule_stats: dict[str, dict] = {}
    for rule in VALIDATION_RULES:
        char_col = None
        for hint in rule["char_hints"]:
            char_col = find_column(df_columns, hint)
            if char_col:
                break
        rule_stats[rule["label"]] = {
            "applicable": char_col is not None,
            "conflicts": 0,
        }

    # Додаємо лічильники для перевірки мов та семантики
    language_conflicts = 0
    semantic_conflicts = 0

    # Обробляємо кожен рядок товару.
    # Примітка: iterrows() зручний для складної рядкової логіки, але для дуже
    # великих фідів (>100k рядків) розгляньте можливість рефакторингу на df.apply().
    for idx, row in df.iterrows():
        row_errors: list[str] = []

        # Перевірка відповідності мов (російська/українська)
        try:
            lang_error_msg, lang_conflict_cols = check_language_mismatch(row, df_columns)
            if lang_error_msg:
                row_errors.append(lang_error_msg)
                language_conflicts += 1
                for col in lang_conflict_cols:
                    conflict_cells[(idx, col)] = True
        except Exception as exc:
            print(
                f"Попередження: помилка при перевірці мови у рядку {idx}: {exc}"
            )

        # Семантична перевірка (суперечності між текстовими полями)
        try:
            sem_error_msg, sem_conflict_cols = check_semantic_conflicts(row, df_columns)
            if sem_error_msg:
                row_errors.append(sem_error_msg)
                semantic_conflicts += 1
                for col in sem_conflict_cols:
                    conflict_cells[(idx, col)] = True
        except Exception as exc:
            print(
                f"Попередження: помилка при семантичній перевірці у рядку {idx}: {exc}"
            )

        for rule in VALIDATION_RULES:
            try:
                error_msg, conflict_cols = check_conflicts(row, df_columns, rule)
            except Exception as exc:
                # Не зупиняємося при помилці в одному правилі — обробляємо далі
                print(
                    f"Попередження: помилка при перевірці '{rule['label']}' "
                    f"у рядку {idx}: {exc}"
                )
                continue

            if error_msg:
                row_errors.append(error_msg)
                rule_stats[rule["label"]]["conflicts"] += 1
                for col in conflict_cols:
                    conflict_cells[(idx, col)] = True

        if row_errors:
            df.at[idx, "Помилки"] = " | ".join(row_errors)
            conflict_cells[(idx, "Помилки")] = True

    # Підраховуємо та виводимо загальні результати
    errors_count = (df["Помилки"] != "").sum()
    print(f"\nЗнайдено конфліктів: {errors_count} рядків із {len(df)}.")

    # Виводимо звіт по перевірці мов
    print("\nПеревірка відповідності мов:")
    if language_conflicts == 0:
        print(f"  ✓  Невідповідностей мов не знайдено")
    else:
        print(f"  ✗  Знайдено {language_conflicts} невідповідностей мов")

    # Виводимо звіт по семантичній перевірці
    print("\nСемантична перевірка:")
    if semantic_conflicts == 0:
        print(f"  ✓  Семантичних суперечностей не знайдено")
    else:
        print(f"  ✗  Знайдено {semantic_conflicts} семантичних суперечностей")

    # Виводимо повний звіт по характеристиках
    print("\nЗвіт по характеристиках:")
    for label, stats in rule_stats.items():
        if not stats["applicable"]:
            print(f"  —  {label}: колонка відсутня у файлі (пропущено)")
        elif stats["conflicts"] == 0:
            print(f"  ✓  {label}: конфліктів не знайдено")
        else:
            print(f"  ✗  {label}: знайдено {stats['conflicts']} конфліктів")

    if conflict_cells:
        print("\nКоординати конфліктних клітинок (рядок, колонка):")
        for row_idx, col_name in sorted(conflict_cells.keys()):
            print(f"  Рядок {row_idx + 2} (Excel), Колонка «{col_name}»")

    # Зберігаємо у Excel із підсвіткою
    _save_with_highlights(df, output_file, conflict_cells)
    print(f"\nРезультат збережено: {output_file}")

    return df


# ===========================================================================
# CLI-точка входу
# ===========================================================================


def main() -> None:
    """Запуск програми з командного рядка.

    Використання:
        python checkr.py <вхідний_файл> [<вихідний_файл>]

    Приклади:
        python checkr.py products.csv result.xlsx
        python checkr.py products.xlsx result.xlsx
        python checkr.py products.csv  # створить products_result.xlsx
    """
    parser = argparse.ArgumentParser(
        prog="checkr",
        description=(
            "Валідація товарного фіду e-commerce: "
            "пошук конфліктів між текстовими полями та характеристиками товару."
        ),
    )
    parser.add_argument("input", help="Вхідний файл (.csv, .xlsx або .xls)")
    parser.add_argument(
        "output",
        nargs="?",
        help="Вихідний файл (.xlsx). Якщо не вказано, створюється <input>_result.xlsx",
    )
    args = parser.parse_args()

    # Якщо output не вказано, створити автоматичне ім'я
    if args.output is None:
        input_path = Path(args.input)
        output_name = f"{input_path.stem}_result.xlsx"
        args.output = str(input_path.parent / output_name)

    try:
        validate_feed(args.input, args.output)
    except FileNotFoundError as exc:
        print(f"Помилка: {exc}", file=sys.stderr)
        sys.exit(1)
    except ValueError as exc:
        print(f"Помилка формату: {exc}", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"Несподівана помилка: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
